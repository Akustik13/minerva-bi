from django.contrib import admin
from core.mixins import AuditableMixin
from crm.utils import sync_customer_from_order
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.db.models import Sum
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import path, reverse
from django import forms
from django.forms import widgets
from .models import SalesOrder, SalesOrderLine, SalesSource
from .forms import SalesImportForm, SalesExcelUploadForm
import openpyxl
from decimal import Decimal, InvalidOperation
import re


COMMON_CURRENCIES = ["USD", "EUR", "GBP", "CHF", "JPY", "CNY", "PLN", "UAH",
                     "CZK", "HUF", "CAD", "AUD", "SEK", "NOK", "DKK"]


class CurrencyComboWidget(widgets.TextInput):
    """Text input with datalist — shows all options on focus, allows free entry."""
    def render(self, name, value, attrs=None, renderer=None):
        if attrs is None:
            attrs = {}
        list_id = "sv-currency-datalist"
        attrs["list"] = list_id
        attrs.setdefault("style", "width:82px")
        # On focus: clear so all datalist options appear; on blur: restore if empty
        attrs["onfocus"] = "this._sv=this.value;this.value=''"
        attrs["onblur"] = "if(!this.value)this.value=this._sv"
        html = super().render(name, value, attrs, renderer)
        options = "".join(f'<option value="{c}">' for c in COMMON_CURRENCIES)
        html += f'<datalist id="{list_id}">{options}</datalist>'
        return mark_safe(html)


class SalesOrderLineInlineForm(forms.ModelForm):
    currency = forms.CharField(
        max_length=8,
        required=False,
        initial="USD",
        widget=CurrencyComboWidget(),
    )

    class Meta:
        model = SalesOrderLine
        fields = "__all__"
import json
import os
import tempfile
import uuid
from datetime import date
from pathlib import Path
import shutil


_SOURCE_TIPS = {
    "slug": (
        "Код (slug) — внутрішній ідентифікатор джерела",
        "Це короткий технічний код, який система зберігає\n"
        "всередині кожного замовлення замість повної назви.\n\n"
        "Правила:\n"
        "• тільки латинські літери та цифри\n"
        "• замість пробілу — підкреслення _\n"
        "• Приклади: digikey, nova_post, webshop\n\n"
        "Структуру бази даних НЕ змінює — це просто\n"
        "текстове значення, яке зберігається в замовленнях.\n\n"
        "❗ Якщо після збереження джерела ви вирішите\n"
        "змінити цей код — замовлення що вже мають старий код\n"
        "перестануть відображатися під цим джерелом.\n"
        "Назву і колір змінювати можна вільно."
    ),
    "name": (
        "Назва — відображається в інтерфейсі",
        "Це те, що бачать користувачі в списках та бейджах.\n"
        "Можна писати будь-якою мовою, з пробілами.\n\n"
        "Змінювати можна вільно — на базу даних не впливає."
    ),
    "color": (
        "Колір бейджу (мітки) у форматі HEX",
        "Визначає колір кольорової мітки джерела замовлення.\n\n"
        "Формат: # і 6 символів (цифри 0-9 та літери a-f)\n"
        "Приклади:\n"
        "• #e91e63 — рожевий (DigiKey)\n"
        "• #ff9800 — помаранчевий (Nova Post)\n"
        "• #4caf50 — зелений\n"
        "• #607d8b — сірий (default)\n\n"
        "Змінювати можна вільно — лише зовнішній вигляд."
    ),
    "order": (
        "Порядок відображення в списках та дропдаунах",
        "Менше число — вище в списку.\n\n"
        "Приклади: 1, 2, 3 … або 10, 20, 30\n"
        "(числа з запасом зручніші — легше вставити між ними).\n\n"
        "Змінювати можна вільно — на дані не впливає."
    ),
}


def _mtip_sales(field_key):
    title, body = _SOURCE_TIPS.get(field_key, ("", ""))
    body_html = body.replace("\n", "<br>")
    return mark_safe(
        f'<i class="mtip" aria-label="{title}">'
        f'?<span class="mtip-body"><b>{title}</b><br>{body_html}</span>'
        f'</i>'
    )


@admin.register(SalesSource)
class SalesSourceAdmin(admin.ModelAdmin):
    list_display       = ("order", "slug", "name", "color_badge")
    list_display_links = ("slug",)
    list_editable      = ("order",)
    search_fields      = ("slug", "name")
    ordering           = ("order", "name")

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        for field_key in ("slug", "name", "color", "order"):
            if field_key in form.base_fields:
                form.base_fields[field_key].help_text = _mtip_sales(field_key)
        return form

    def color_badge(self, obj):
        return format_html(
            '<span style="background:{};color:#fff;padding:3px 12px;border-radius:10px;'
            'font-size:12px;font-weight:bold">{}</span>',
            obj.color, obj.name)
    color_badge.short_description = "Вигляд бейджу"


class OrderPackagingInline(admin.TabularInline):
    model = None  # встановлюється нижче після імпорту
    extra = 0
    fields = ('packaging', 'qty_boxes', 'actual_weight_g', 'notes')
    verbose_name        = 'Тип коробки'
    verbose_name_plural = '📦 Фактична упаковка (статистика)'

    def get_formset(self, request, obj=None, **kwargs):
        fs = super().get_formset(request, obj, **kwargs)
        fs.form.base_fields['qty_boxes'].label = 'К-сть шт (однакових)'
        fs.form.base_fields['qty_boxes'].help_text = 'Скільки однакових коробок цього розміру (наприклад 5 = п\'ять однакових коробок)'
        return fs

try:
    from shipping.models import OrderPackaging as _OP
    OrderPackagingInline.model = _OP
except Exception:
    OrderPackagingInline = None  # shipping не встановлено


class SalesOrderLineInline(admin.TabularInline):
    model = SalesOrderLine
    form = SalesOrderLineInlineForm
    extra = 0
    readonly_fields = ("stock_status",)
    fields = ("product", "sku_raw", "qty", "unit_price", "total_price", "currency", "stock_status")
    autocomplete_fields = ("product",)

    # Темний фон для inline


    def stock_status(self, obj):
        if not obj.product:
            if obj.sku_raw:
                return format_html('<span style="color:#ff9800">⚠️ {} не в базі</span>', obj.sku_raw)
            return "—"
        from inventory.models import InventoryTransaction
        result = InventoryTransaction.objects.filter(product=obj.product).aggregate(total=Sum('qty'))
        stock = float(result['total'] or 0)
        needed = float(obj.qty or 0)
        if stock <= 0:
            return format_html('<span style="background:#f44336;color:#fff;padding:2px 8px;border-radius:8px;font-size:11px">🚫 Немає (0)</span>')
        elif stock < needed:
            return format_html('<span style="background:#ff9800;color:#fff;padding:2px 8px;border-radius:8px;font-size:11px">⚠️ {} / {}</span>', int(stock), int(needed))
        return format_html('<span style="background:#4caf50;color:#fff;padding:2px 8px;border-radius:8px;font-size:11px">✅ {} шт.</span>', int(stock))
    stock_status.short_description = "На складі"


class UnshippedFilter(admin.SimpleListFilter):
    title = "Відправлено"
    parameter_name = "shipped"
    def lookups(self, request, model_admin):
        return [("no", "⏳ Не відправлено"), ("yes", "✅ Відправлено")]
    def queryset(self, request, queryset):
        if self.value() == "no": return queryset.filter(shipped_at__isnull=True)
        if self.value() == "yes": return queryset.filter(shipped_at__isnull=False)
        return queryset


class OverdueFilter(admin.SimpleListFilter):
    title = "Прострочені"
    parameter_name = "overdue"
    def lookups(self, request, model_admin):
        return [("1", "🔴 Прострочений дедлайн")]
    def queryset(self, request, queryset):
        if self.value() == "1":
            from datetime import date
            return queryset.filter(
                shipping_deadline__lt=date.today(),
                status__in=["received", "processing"],
            )
        return queryset


class ShippedThisMonthFilter(admin.SimpleListFilter):
    title = "Відправлено"
    parameter_name = "shipped_period"
    def lookups(self, request, model_admin):
        return [("month", "🚚 Цього місяця")]
    def queryset(self, request, queryset):
        if self.value() == "month":
            from datetime import date
            today = date.today()
            return queryset.filter(shipped_at__year=today.year, shipped_at__month=today.month)
        return queryset


def export_sales_excel(modeladmin, request, queryset):
    try:
        from exports import export_sales
        return export_sales(queryset)
    except Exception as e:
        from django.contrib import messages
        messages.error(request, f"Помилка: {e}")
export_sales_excel.short_description = "📥 Експортувати в Excel"


class SalesOrderDocumentsForm(forms.ModelForm):
    """Форма з полями для завантаження документів"""
    
    # Видаліть upload_documents взагалі - будемо обробляти через request.FILES
    
    local_base_path = forms.CharField(
        required=False,
        max_length=500,
        widget=forms.HiddenInput(),
    )
    
    class Meta:
        model = SalesOrder
        fields = '__all__'

@admin.register(SalesOrder)
class SalesOrderAdmin(AuditableMixin, admin.ModelAdmin):
    form = SalesOrderDocumentsForm
    # ── Прибрано: phone, tracking_number, customer_link (без лінку) ───────────
    list_display = (
        "order_number", "source_badge", "status_badge", "order_date_fmt", 'deadline_display',
        "customer_link_display", "country_display",
        "shipped_badge",
        "items_count", "items_summary", "order_total",
        "stock_warning",
        "label_buttons_list",
        # "customer_link",  # disabled - no FK
    )
    search_fields = ("order_number", "tracking_number", "client", "email", "phone",
                     "addr_city", "addr_street")
    list_filter   = ("source", "status", "addr_country", UnshippedFilter, OverdueFilter, ShippedThisMonthFilter)
    ordering       = ("-order_date",)
    date_hierarchy = "order_date"
    actions        = [
        export_sales_excel,
        # Статус
        "action_set_status_received",
        "action_set_status_processing",
        "action_set_status_shipped",
        "action_set_status_cancelled",
        # Валюта продажу
        "action_set_currency_usd",
        "action_set_currency_eur",
        "action_set_currency_gbp",
        # Валюта доставки
        "action_set_shipping_currency_usd",
        "action_set_shipping_currency_eur",
        "action_set_shipping_currency_gbp",
        # Тип документа
        "action_set_doctype_sale",
        "action_set_doctype_sample",
        "action_set_doctype_transfer",
        "action_set_doctype_warranty",
        "action_set_doctype_other",
        # Впливає на склад
        "action_set_affects_stock_yes",
        "action_set_affects_stock_no",
        # Джерело (проміжна форма)
        "action_change_source",
    ]
    preserve_filters = True
    inlines        = [i for i in [SalesOrderLineInline, OrderPackagingInline] if i is not None]

    # ── Actions: Статус ───────────────────────────────────────────────────────

    def action_set_status_received(self, request, queryset):
        n = queryset.update(status="received")
        self.message_user(request, f"📥 Статус «Отримано» встановлено для {n} замовлень.")
    action_set_status_received.short_description = "📥 Статус → Отримано"

    def action_set_status_processing(self, request, queryset):
        n = queryset.update(status="processing")
        self.message_user(request, f"⚙️ Статус «В обробці» встановлено для {n} замовлень.")
    action_set_status_processing.short_description = "⚙️ Статус → В обробці"

    def action_set_status_shipped(self, request, queryset):
        n = queryset.update(status="shipped")
        self.message_user(request, f"🚀 Статус «Відправлено» встановлено для {n} замовлень.")
    action_set_status_shipped.short_description = "🚀 Статус → Відправлено"

    def action_set_status_cancelled(self, request, queryset):
        n = queryset.update(status="cancelled")
        self.message_user(request, f"❌ Статус «Скасовано» встановлено для {n} замовлень.")
    action_set_status_cancelled.short_description = "❌ Статус → Скасовано"

    # ── Actions: Валюта продажу ───────────────────────────────────────────────

    def action_set_currency_usd(self, request, queryset):
        n = queryset.update(currency="USD")
        self.message_user(request, f"💵 Валюта продажу → USD для {n} замовлень.")
    action_set_currency_usd.short_description = "💵 Валюта продажу → USD"

    def action_set_currency_eur(self, request, queryset):
        n = queryset.update(currency="EUR")
        self.message_user(request, f"💶 Валюта продажу → EUR для {n} замовлень.")
    action_set_currency_eur.short_description = "💶 Валюта продажу → EUR"

    def action_set_currency_gbp(self, request, queryset):
        n = queryset.update(currency="GBP")
        self.message_user(request, f"💷 Валюта продажу → GBP для {n} замовлень.")
    action_set_currency_gbp.short_description = "💷 Валюта продажу → GBP"

    # ── Actions: Валюта доставки ──────────────────────────────────────────────

    def action_set_shipping_currency_usd(self, request, queryset):
        n = queryset.update(shipping_currency="USD")
        self.message_user(request, f"💵 Валюта доставки → USD для {n} замовлень.")
    action_set_shipping_currency_usd.short_description = "💵 Валюта доставки → USD"

    def action_set_shipping_currency_eur(self, request, queryset):
        n = queryset.update(shipping_currency="EUR")
        self.message_user(request, f"💶 Валюта доставки → EUR для {n} замовлень.")
    action_set_shipping_currency_eur.short_description = "💶 Валюта доставки → EUR"

    def action_set_shipping_currency_gbp(self, request, queryset):
        n = queryset.update(shipping_currency="GBP")
        self.message_user(request, f"💷 Валюта доставки → GBP для {n} замовлень.")
    action_set_shipping_currency_gbp.short_description = "💷 Валюта доставки → GBP"

    # ── Actions: Тип документа ────────────────────────────────────────────────

    def action_set_doctype_sale(self, request, queryset):
        n = queryset.update(document_type="SALE")
        self.message_user(request, f"📄 Тип → Sale для {n} замовлень.")
    action_set_doctype_sale.short_description = "📄 Тип документа → Sale"

    def action_set_doctype_sample(self, request, queryset):
        n = queryset.update(document_type="SAMPLE")
        self.message_user(request, f"🧪 Тип → Sample для {n} замовлень.")
    action_set_doctype_sample.short_description = "🧪 Тип документа → Sample"

    def action_set_doctype_transfer(self, request, queryset):
        n = queryset.update(document_type="TRANSFER")
        self.message_user(request, f"↔️ Тип → Transfer для {n} замовлень.")
    action_set_doctype_transfer.short_description = "↔️ Тип документа → Transfer"

    def action_set_doctype_warranty(self, request, queryset):
        n = queryset.update(document_type="WARRANTY")
        self.message_user(request, f"🛡️ Тип → Warranty для {n} замовлень.")
    action_set_doctype_warranty.short_description = "🛡️ Тип документа → Warranty"

    def action_set_doctype_other(self, request, queryset):
        n = queryset.update(document_type="OTHER")
        self.message_user(request, f"📎 Тип → Other для {n} замовлень.")
    action_set_doctype_other.short_description = "📎 Тип документа → Other"

    # ── Actions: Впливає на склад ─────────────────────────────────────────────

    def action_set_affects_stock_yes(self, request, queryset):
        n = queryset.update(affects_stock=True)
        self.message_user(request, f"✅ «Впливає на склад» увімкнено для {n} замовлень.")
    action_set_affects_stock_yes.short_description = "✅ Склад → Впливає"

    def action_set_affects_stock_no(self, request, queryset):
        n = queryset.update(affects_stock=False)
        self.message_user(request, f"⬜ «Впливає на склад» вимкнено для {n} замовлень.")
    action_set_affects_stock_no.short_description = "⬜ Склад → Не впливає"

    # ── Action: Змінити джерело (проміжна форма) ──────────────────────────────

    def action_change_source(self, request, queryset):
        from .models import SalesSource
        from django.contrib.admin import helpers

        if "apply" in request.POST:
            slug = request.POST.get("source_slug", "").strip()
            if slug:
                n = queryset.update(source=slug)
                self.message_user(request, f"📦 Джерело «{slug}» встановлено для {n} замовлень.")
            return None

        sources = SalesSource.objects.all()
        return render(
            request,
            "admin/sales/action_change_source.html",
            {
                "queryset": queryset,
                "sources": sources,
                "action_checkbox_name": helpers.ACTION_CHECKBOX_NAME,
                **self.admin_site.each_context(request),
            },
        )
    action_change_source.short_description = "📦 Змінити джерело..."

    fieldsets = (
        ("📦 Замовлення", {
            "fields": ("source", "status", "document_type", "affects_stock",
                       "order_number", "order_date")
        }),
        ("👤 Клієнт", {
            "fields": ("client", "contact_name", "email", "phone")
        }),
        ("🚚 Доставка", {
            "fields": (
                ("addr_street",),
                ("addr_city", "addr_zip", "addr_state", "addr_country"),
                "shipping_deadline", "shipped_at", "delivered_at",
                "shipping_courier", "tracking_number", "lieferschein_nr",
                ("shipping_cost", "shipping_currency"),
            )
        }),
        ("📋 Legacy адреса (raw)", {
            "fields": ("shipping_region", "shipping_address"),
            "classes": ("collapse",),
            "description": "Оригінальний текстовий формат адреси — збережено для сумісності з імпортом",
        }),
        
        ("📦 Залишки на складі", {
            "fields": ("stock_summary",)
        }),
        ("🏷️ Етикетки DYMO", {
            "fields": ("label_buttons_detail", "label_upload_widget"),
            "description": "Натисніть кнопку — файл завантажиться і відкриється в DYMO Label Software",
        }),
        
        ("📦 Пакування", {
            "fields": ("packaging_panel",),
            "classes": ("collapse",),
            "description": "Рекомендована упаковка на основі товарів замовлення. Фактичну упаковку зафіксуй нижче (inline).",
        }),
        ("📄 Автоматичні документи", {
            "fields": ("doc_buttons",),
            "description": "PDF генерується на льоту з даних замовлення",
        }),
        ("📄 Документи замовлення", {
            "fields": ("documents_list", "upload_widget"),
            "classes": ("collapse",),
            "description": "Завантаження документів: етикетки, декларації, чеки тощо. Автоматичне збереження на сервер та локально."
        }),
    )

    readonly_fields = ['stock_summary', 'label_buttons_detail', 'label_upload_widget',
                       'documents_list', 'upload_widget', 'doc_buttons', 'packaging_panel']
    
    def _docs_panel_html(self, obj):
        """Inner HTML for the documents panel (used by documents_list and doc_list_view)."""
        from django.conf import settings
        media_path = settings.MEDIA_ROOT / 'orders' / obj.source / obj.order_number

        if not media_path.exists():
            return '<em style="color:#999">Документів немає</em>'

        files = list(media_path.glob('*'))
        if not files:
            return '<em style="color:#999">Документів немає</em>'

        delete_url = f'/admin/sales/salesorder/{obj.pk}/doc/delete/'
        rows = ''
        for file in sorted(files):
            size_kb = file.stat().st_size / 1024
            url = f'{settings.MEDIA_URL}orders/{obj.source}/{obj.order_number}/{file.name}'
            rows += (
                f'<tr style="border-bottom:1px solid #2a3f52">'
                f'<td style="padding:8px 4px 8px 0"><span style="color:#e0e0e0">📄 {file.name}</span></td>'
                f'<td style="padding:8px 4px;text-align:right;color:#9aafbe;font-size:11px;white-space:nowrap">'
                f'{size_kb:.1f} KB</td>'
                f'<td style="padding:8px 0;text-align:right;white-space:nowrap">'
                f'<a href="{url}" target="_blank" '
                f'style="background:#417690;color:#fff;padding:5px 10px;text-decoration:none;'
                f'border-radius:3px;font-size:11px;margin-right:4px">⬇️</a>'
                f'<button type="button"'
                f' data-del-url="{delete_url}"'
                f' data-filename="{file.name}"'
                f' onclick="deleteOrderDoc(this)"'
                f' style="background:#b71c1c;color:#fff;border:none;padding:5px 10px;'
                f'border-radius:3px;font-size:11px;cursor:pointer"'
                f' title="Видалити файл з сервера">🗑️</button>'
                f'</td></tr>'
            )

        return (
            f'<div style="color:#7ab3cc;font-weight:bold;margin-bottom:10px;font-size:13px">'
            f'📁 Завантажені документи ({len(files)})</div>'
            f'<table style="width:100%;border-collapse:collapse">{rows}</table>'
        )

    def documents_list(self, obj):
        """Показує список завантажених документів з кнопками видалення."""
        if not obj.pk:
            return format_html('<em style="color:#999">Збережіть замовлення для завантаження документів</em>')

        js = (
            '<script>'
            'if(!window._delDocDef){'
            'window._delDocDef=true;'
            'function _refreshDocsPanel(pk){'
            'fetch("/admin/sales/salesorder/"+pk+"/doc/list/",'
            '{headers:{"X-Requested-With":"XMLHttpRequest"}})'
            '.then(r=>r.text())'
            '.then(html=>{'
            'var p=document.getElementById("order-docs-panel");'
            'if(p)p.innerHTML=html;'
            '})'
            '.catch(()=>{})'
            '}'
            'function deleteOrderDoc(btn){'
            'var fn=btn.dataset.filename;'
            'if(!confirm("Видалити файл \\""+fn+"\\" з сервера?"))return;'
            'btn.disabled=true;btn.textContent="…";'
            'var csrf=document.querySelector("[name=csrfmiddlewaretoken]").value;'
            'var fd=new FormData();fd.append("filename",fn);'
            'var pk=document.getElementById("order-docs-panel").dataset.pk;'
            'fetch(btn.dataset.delUrl,{method:"POST",'
            'headers:{"X-CSRFToken":csrf},body:fd})'
            '.then(r=>r.json())'
            '.then(d=>{'
            'if(d.ok){_refreshDocsPanel(pk);}'
            'else{alert("Помилка: "+(d.error||"?"));'
            'btn.disabled=false;btn.textContent="🗑️";}})'
            '.catch(()=>{alert("Помилка мережі");'
            'btn.disabled=false;btn.textContent="🗑️";})'
            '}}'
            '</script>'
        )

        inner = self._docs_panel_html(obj)
        return format_html(
            '{}<div id="order-docs-panel" data-pk="{}"'
            ' style="background:#1e2a35;border:1px solid #2a3f52;border-radius:4px;'
            'padding:15px;margin-top:10px">'
            '{}</div>',
            mark_safe(js), obj.pk, mark_safe(inner)
        )

    documents_list.short_description = '📋 Завантажені документи'

    def doc_buttons(self, obj):
        """PDF generation buttons with inline quick-edit override panels."""
        if not obj.pk:
            return mark_safe('<em style="color:#607d8b">Збережіть замовлення</em>')

        # Load DocumentSettings for pre-populating override fields
        try:
            from config.models import DocumentSettings, DOCUMENT_LANGUAGE_CHOICES, CN23_TYPE_CHOICES
            ds = DocumentSettings.get()
            lang_opts = ''.join(
                f'<option value="{v}"{"selected" if ds.doc_language == v else ""}>{lbl}</option>'
                for v, lbl in DOCUMENT_LANGUAGE_CHOICES
            )
            cn_opts = ''.join(
                f'<option value="{v}"{"selected" if ds.customs_default_type == v else ""}>{lbl}</option>'
                for v, lbl in CN23_TYPE_CHOICES
            )
            show_prices_ck  = 'checked' if ds.packing_list_show_prices else ''
            footer_note_v   = (ds.packing_list_footer_note or '').replace('"', '&quot;')
            pay_terms_v     = (ds.proforma_payment_terms or '').replace('"', '&quot;')
            pf_notes_v      = (ds.proforma_notes or '').replace('"', '&quot;')
            reason_v        = (ds.customs_reason or '').replace('"', '&quot;')
        except Exception:
            lang_opts = '<option value="en">English (EN)</option>'
            cn_opts   = '<option value="SALE">Sale / Verkauf — продаж</option>'
            show_prices_ck = footer_note_v = pay_terms_v = pf_notes_v = reason_v = ''

        # Shared inline styles for override panel inputs
        _inp = ('background:#111c26;border:1px solid #2a3f52;color:#c9d8e4;'
                'padding:5px 8px;border-radius:4px;font-size:12px;'
                'width:100%;box-sizing:border-box;margin-top:3px')
        _sel = ('background:#111c26;border:1px solid #2a3f52;color:#c9d8e4;'
                'padding:5px 8px;border-radius:4px;font-size:12px;margin-top:3px')
        _lbl = 'font-size:11px;color:#9aafbe;display:block'

        pl_panel = (
            f'<div style="display:grid;grid-template-columns:1fr auto;gap:12px;margin-bottom:10px">'
            f'  <div><label style="{_lbl}">🌐 Мова</label>'
            f'  <select name="language" style="{_sel}">{lang_opts}</select></div>'
            f'  <div style="display:flex;align-items:flex-end;gap:6px;padding-bottom:5px">'
            f'    <input type="checkbox" name="show_prices" id="ck-pl-prices" {show_prices_ck}'
            f'     style="width:15px;height:15px;cursor:pointer;margin:0">'
            f'    <label for="ck-pl-prices" style="font-size:12px;color:#c9d8e4;cursor:pointer;white-space:nowrap">💰 Ціни</label>'
            f'  </div>'
            f'</div>'
            f'<div><label style="{_lbl}">📝 Нотатка внизу</label>'
            f'<input type="text" name="footer_note" value="{footer_note_v}" style="{_inp}"'
            f' placeholder="For customs use only…"></div>'
        )

        pf_panel = (
            f'<div style="margin-bottom:10px"><label style="{_lbl}">💳 Умови оплати</label>'
            f'<input type="text" name="payment_terms" value="{pay_terms_v}" style="{_inp}"></div>'
            f'<div><label style="{_lbl}">📝 Примітки</label>'
            f'<input type="text" name="notes" value="{pf_notes_v}" style="{_inp}" placeholder="…"></div>'
        )

        cn_panel = (
            f'<div style="margin-bottom:10px"><label style="{_lbl}">📋 Тип декларації</label>'
            f'<select name="declaration_type" style="{_sel};width:100%">{cn_opts}</select></div>'
            f'<div><label style="{_lbl}">📝 Опис товару (fallback)</label>'
            f'<input type="text" name="reason" value="{reason_v}" style="{_inp}"></div>'
        )

        base = f"/admin/sales/salesorder/{obj.pk}/doc"
        docs = [
            (f"{base}/packing-list/", f"{base}/packing-list/?action=save",
             "📋 Пакувальний лист", "#1976d2", "pl", pl_panel),
            (f"{base}/proforma/",     f"{base}/proforma/?action=save",
             "📄 Proforma Invoice",   "#388e3c", "pf", pf_panel),
            (f"{base}/customs/",      f"{base}/customs/?action=save",
             "🛃 Митна декларація",   "#6a1b9a", "cn", cn_panel),
        ]

        parts = []
        for dl_url, save_url, label, color, doc_id, panel_html in docs:
            parts.append(
                f'<div style="margin:4px 0 10px 0">'
                # ── button row ──────────────────────────────────────────────
                f'<span style="display:inline-flex;align-items:stretch;'
                f'border-radius:7px;overflow:hidden;margin-right:6px">'
                f'<a id="dlbtn-{doc_id}" href="{dl_url}" target="_blank"'
                f' onclick="return _docOpen(this,\'{doc_id}\')"'
                f' style="background:{color};color:#fff;padding:7px 14px;'
                f'text-decoration:none;font-size:13px;font-weight:600;white-space:nowrap">'
                f'{label}</a>'
                f'<button type="button" id="savebtn-{doc_id}"'
                f' data-save-url="{save_url}" data-doc-id="{doc_id}"'
                f' onclick="_docSave(this)"'
                f' title="Зберегти PDF на сервер в папку замовлення"'
                f' style="background:{color};color:#fff;border:none;'
                f'border-left:1px solid rgba(255,255,255,.3);'
                f'padding:7px 11px;cursor:pointer;font-size:13px">💾</button>'
                f'</span>'
                # ── quick-edit toggle ────────────────────────────────────────
                f'<button type="button" id="ovr-toggle-{doc_id}"'
                f' onclick="_toggleDocOvr(\'{doc_id}\')"'
                f' title="Швидке редагування документа"'
                f' style="background:#1e2d3e;color:#9aafbe;border:1px solid #2a3f52;'
                f'border-radius:6px;padding:5px 10px;cursor:pointer;font-size:13px;'
                f'vertical-align:middle;transition:all .15s">⚙️</button>'
                # ── override panel ───────────────────────────────────────────
                f'<div id="doc-ovr-{doc_id}"'
                f' style="display:none;margin-top:8px;padding:13px 15px;'
                f'background:#0b1520;border:1px solid #2a3f52;'
                f'border-left:3px solid {color};border-radius:6px;max-width:480px">'
                f'<div style="font-size:11px;color:#607d8b;margin-bottom:10px;'
                f'letter-spacing:.04em;text-transform:uppercase;font-weight:600">'
                f'⚙️ Редагування для цього документу</div>'
                f'{panel_html}'
                f'</div>'
                f'</div>'
            )

        js = (
            '<script>'
            'if(!window._docOvrDef){window._docOvrDef=true;'

            'function _docParams(id){'
            'var p=document.getElementById("doc-ovr-"+id);'
            'if(!p||p.style.display==="none")return "";'
            'var a=[];'
            'p.querySelectorAll("input[name],select[name],textarea[name]").forEach(function(el){'
            'if(el.type==="checkbox"){a.push(el.name+"="+(el.checked?"1":"0"));}'
            'else{a.push(encodeURIComponent(el.name)+"="+encodeURIComponent(el.value));}'
            '});'
            'return a.join("&");}'

            'function _docOpen(anchor,id){'
            'var extra=_docParams(id);'
            'if(extra){var base=anchor.href.split("?")[0];window.open(base+"?"+extra,"_blank");return false;}'
            'return true;}'

            'function _docSave(btn){'
            'var id=btn.dataset.docId;'
            'var extra=_docParams(id);'
            'var base=btn.dataset.saveUrl.split("?")[0];'
            'var url=base+"?action=save"+(extra?"&"+extra:"");'
            'btn.disabled=true;btn.textContent="⏳";'
            'fetch(url,{headers:{"X-Requested-With":"XMLHttpRequest"}})'
            '.then(r=>r.json())'
            '.then(d=>{'
            'if(d.ok){'
            'btn.textContent="✅";'
            'var pk=document.getElementById("order-docs-panel")?.dataset.pk;'
            'if(pk&&typeof _refreshDocsPanel==="function")_refreshDocsPanel(pk);'
            'var _p=document.getElementById("order-docs-panel");'
            'if(_p){var _fs=_p.closest("fieldset.collapse,fieldset.collapsed");'
            'if(_fs){'
            '_fs.classList.remove("collapsed");_fs.style.display="";'
            'var _rows=_fs.querySelectorAll(".form-row,.form-group");'
            '_rows.forEach(function(r){r.style.display="";});'
            'setTimeout(function(){_fs.scrollIntoView({behavior:"smooth",block:"nearest"});},50);'
            '}}'
            'setTimeout(()=>{btn.disabled=false;btn.textContent="💾";},3000);'
            '}else{'
            'btn.textContent="❌";alert("Помилка: "+(d.error||"?"));'
            'setTimeout(()=>{btn.disabled=false;btn.textContent="💾";},3000);'
            '}})'
            '.catch(()=>{'
            'btn.textContent="❌";alert("Помилка мережі");'
            'setTimeout(()=>{btn.disabled=false;btn.textContent="💾";},3000);'
            '});}'

            'function _toggleDocOvr(id){'
            'var p=document.getElementById("doc-ovr-"+id);'
            'var btn=document.getElementById("ovr-toggle-"+id);'
            'var open=p.style.display!=="none";'
            'p.style.display=open?"none":"block";'
            'btn.style.background=open?"#1e2d3e":"#1a3a5c";'
            'btn.style.color=open?"#9aafbe":"#58a6ff";'
            'btn.style.borderColor=open?"#2a3f52":"#58a6ff";}'

            '}'
            '</script>'
        )
        return mark_safe(js + ''.join(parts))

    doc_buttons.short_description = "📄 Генерація документів"

    def packaging_panel(self, obj):
        """Рекомендована упаковка на основі товарів замовлення."""
        if not obj.pk:
            return mark_safe('<em style="color:#607d8b">Збережіть замовлення</em>')

        lines = list(obj.lines.select_related('product').all())
        if not lines:
            return mark_safe('<span style="color:#607d8b">— рядків замовлення немає</span>')

        try:
            from shipping.models import ProductPackaging
        except ImportError:
            return mark_safe('<span style="color:#607d8b">— модуль shipping недоступний</span>')

        rows_html = []
        total_weight_g = 0
        has_missing_weight = False
        best_box = None

        for line in lines:
            product = line.product
            if not product:
                rows_html.append(
                    f'<tr><td style="color:#9aafbe">{line.sku_raw or "?"}</td>'
                    f'<td style="color:#9aafbe">{line.qty}</td>'
                    f'<td style="color:#607d8b">— товар не в базі</td>'
                    f'<td style="color:#607d8b">—</td></tr>'
                )
                continue

            rec = ProductPackaging.objects.filter(
                product=product, is_default=True
            ).select_related('packaging').first()

            if rec:
                box_str = str(rec.packaging)
                boxes_needed = max(1, int(-(-float(line.qty) // rec.qty_per_box)))  # ceil
                if rec.estimated_weight_g:
                    line_weight = rec.estimated_weight_g * boxes_needed
                    total_weight_g += line_weight
                    weight_str = f'{line_weight} г'
                    if best_box is None:
                        best_box = rec.packaging
                else:
                    weight_str = '<span style="color:#ff9800">⚠️ не вказана</span>'
                    has_missing_weight = True
                rows_html.append(
                    f'<tr>'
                    f'<td style="color:#9aafbe;font-family:monospace">{product.sku}</td>'
                    f'<td style="color:#c9d8e4;text-align:center">{line.qty}</td>'
                    f'<td style="color:#80cbc4">{box_str} × {boxes_needed}</td>'
                    f'<td style="color:#aed581">{weight_str}</td>'
                    f'</tr>'
                )
            else:
                rows_html.append(
                    f'<tr>'
                    f'<td style="color:#9aafbe;font-family:monospace">{product.sku}</td>'
                    f'<td style="color:#c9d8e4;text-align:center">{line.qty}</td>'
                    f'<td style="color:#607d8b">— упаковку не налаштовано '
                    f'<a href="/admin/inventory/product/{product.pk}/change/#packaging" '
                    f'target="_blank" style="color:#2196f3">налаштувати</a></td>'
                    f'<td style="color:#607d8b">—</td>'
                    f'</tr>'
                )

        table = (
            '<table style="width:100%;border-collapse:collapse;font-size:12px;margin-bottom:10px">'
            '<thead><tr style="border-bottom:1px solid #2a3f52">'
            '<th style="text-align:left;padding:4px 8px;color:#607d8b">SKU</th>'
            '<th style="text-align:center;padding:4px 8px;color:#607d8b">К-сть</th>'
            '<th style="text-align:left;padding:4px 8px;color:#607d8b">Упаковка</th>'
            '<th style="text-align:left;padding:4px 8px;color:#607d8b">Вага</th>'
            '</tr></thead><tbody>'
            + ''.join(rows_html)
            + '</tbody></table>'
        )

        # Підсумок
        if total_weight_g > 0:
            weight_kg = total_weight_g / 1000
            summary = (
                f'<div style="padding:8px 12px;background:#0d2137;border-radius:6px;'
                f'border-left:3px solid #2196f3;font-size:12px;color:#9aafbe">'
                f'<b style="color:#c9d8e4">Орієнтовна вага посилки:</b> '
                f'<span style="color:#aed581;font-weight:700">{weight_kg:.3f} кг ({total_weight_g} г)</span>'
            )
            if best_box:
                summary += (
                    f' &nbsp;|&nbsp; <b style="color:#c9d8e4">Рекомендована коробка:</b> '
                    f'<span style="color:#80cbc4">{best_box}</span>'
                )
            if has_missing_weight:
                summary += '<br><span style="color:#ff9800">⚠️ Деякі товари без вказаної ваги — результат неточний</span>'
            summary += '</div>'
        else:
            summary = '<div style="color:#607d8b;font-size:12px">Заповни поле «Вага нетто (г/шт)» у картці товару для розрахунку</div>'

        # Зафіксована упаковка
        fixed = ''
        try:
            from shipping.models import OrderPackaging
            used = OrderPackaging.objects.filter(order=obj).select_related('packaging')
            if used.exists():
                fixed_items = ', '.join(
                    f'{u.packaging} ×{u.qty_boxes}'
                    + (f' ({u.actual_weight_g} г)' if u.actual_weight_g else '')
                    for u in used
                )
                fixed = (
                    f'<div style="margin-top:8px;padding:6px 12px;background:#0d2b1a;'
                    f'border-radius:6px;border-left:3px solid #4caf50;font-size:12px">'
                    f'<span style="color:#4caf50;font-weight:700">✅ Зафіксовано:</span> '
                    f'<span style="color:#9aafbe">{fixed_items}</span></div>'
                )
        except Exception:
            pass

        return mark_safe(table + summary + fixed)

    packaging_panel.short_description = '📦 Рекомендована упаковка'

    def upload_widget(self, obj):
        """HTML-віджет для завантаження файлів. JS — у salesorder_change_form.html."""
        if not obj.pk:
            return format_html('<em style="color:#7d8590">Збережіть замовлення спочатку</em>')

        upload_url = str(reverse('admin:sales_salesorder_upload_docs', args=[obj.pk]))
        browse_url = str(reverse('admin:sales_salesorder_browse_dir'))

        # Тільки HTML + data-атрибути. Весь JS — у change_form_template (salesorder_change_form.html).
        # Це гарантує виконання скрипту незалежно від collapsed-fieldset.
        # Лише HTML з data-атрибутами. Скрипт — у salesorder_change_form.html → block extrahead.
        html = (
            '<div id="docWidgetRoot"'
            ' data-browse-url="' + browse_url + '"'
            ' data-upload-url="' + upload_url + '"'
            ' data-order-number="' + (obj.order_number or '') + '"'
            ' data-source="' + (obj.source or '') + '"'
            ' style="background:rgba(255,255,255,0.05);border:1px solid rgba(255,255,255,0.1);'
            'padding:16px;border-radius:8px">'
            '<p style="margin:0 0 8px;font-weight:bold;opacity:0.9">📤 Оберіть файли для завантаження:</p>'
            '<input type="file" id="docFiles" multiple style="margin-bottom:10px;display:block;opacity:0.85">'
            '<p style="margin:0 0 8px;font-size:12px;color:#888">'
            'Можна обрати кілька файлів одночасно (етикетки, декларації, чеки тощо)</p>'
            '<div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;flex-wrap:wrap">'
            '<button type="button" id="docBrowseBtn"'
            ' style="background:#37474f;color:#b0bec5;border:1px solid #546e7a;'
            'padding:6px 14px;border-radius:6px;cursor:pointer;font-size:13px">'
            '📂 Вибрати папку на ПК</button>'
            '<span id="docSelectedPath" style="font-size:12px;color:#4caf50;display:none">'
            '✅ <span id="docPathDisplay" style="font-family:monospace;word-break:break-all"></span></span>'
            '<span id="docPathHint" style="font-size:11px;color:#546e7a;display:none">'
            '🕓 <span id="docPathHintText"></span></span>'
            '</div>'
            '<p style="margin:0 0 10px;font-size:11px;color:#546e7a">'
            '💡 <b style="color:#607d8b">Chrome / Edge</b>: відкриється нативний діалог Windows — '
            'вибираєте папку як у провіднику, файли записуються напряму на ПК.<br>'
            '🔸 Firefox: відкриється навігатор по серверу (вкажіть шлях вручну або '
            'примонтуйте каталог у docker-compose).</p>'
            '<button type="button" id="docUploadBtn"'
            ' style="background:#417690;color:#fff;border:none;padding:8px 20px;'
            'border-radius:6px;cursor:pointer;font-size:14px;font-weight:bold">✅ Завантажити документи</button>'
            '<div id="docResult" style="margin-top:10px;font-size:13px"></div>'
            '</div>'
            # ── Modal: folder browser ──
            '<div id="dirBrowserModal"'
            ' style="display:none;position:fixed;top:0;left:0;width:100%;height:100%;'
            'background:rgba(0,0,0,.72);z-index:100000">'
            '<div style="position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);'
            'background:#1e2a35;border:1px solid #2a3f52;border-radius:10px;'
            'padding:24px;width:660px;max-width:94vw;'
            'display:flex;flex-direction:column;box-shadow:0 8px 40px rgba(0,0,0,.7)">'
            # header
            '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:14px">'
            '<span style="color:#c9d8e4;font-weight:700;font-size:15px">📂 Вибрати папку</span>'
            '<button type="button" id="dirBrowserCloseBtn"'
            ' style="background:none;border:none;color:#9aafbe;font-size:22px;'
            'cursor:pointer;line-height:1;padding:0">✕</button>'
            '</div>'
            # editable path bar
            '<div style="display:flex;gap:6px;margin-bottom:8px">'
            '<input type="text" id="dirBrowserPathInput"'
            ' placeholder="Введіть шлях або оберіть нижче…"'
            ' style="flex:1;background:#111c26;border:1px solid #2a3f52;border-radius:4px;'
            'padding:7px 10px;font-size:12px;font-family:monospace;color:#c9d8e4;outline:none">'
            '<button type="button" id="dirBrowserGoBtn"'
            ' style="background:#2a3f52;color:#9aafbe;border:1px solid #3a5570;'
            'border-radius:4px;padding:7px 14px;cursor:pointer;font-size:13px;white-space:nowrap">'
            '→ Перейти</button>'
            '</div>'
            # info note
            '<div style="font-size:11px;color:#546e7a;margin-bottom:10px;line-height:1.5">'
            'ℹ️ Браузер показує файлову систему <b style="color:#607d8b">сервера</b>. '
            'Для Django у Docker — вкажіть Linux-шлях до примонтованого каталогу. '
            'Для локального запуску на Windows — шлях у форматі <code style="color:#9aafbe">C:\\Users\\...</code>'
            '</div>'
            # dir list
            '<div id="dirBrowserList"'
            ' style="overflow-y:auto;border:1px solid #2a3f52;border-radius:4px;'
            'background:#0d1117;min-height:160px;max-height:320px"></div>'
            # footer
            '<div style="margin-top:16px;display:flex;gap:10px;justify-content:flex-end">'
            '<button type="button" id="dirBrowserCancelBtn"'
            ' style="background:#37474f;color:#b0bec5;border:1px solid #546e7a;'
            'padding:8px 18px;border-radius:6px;cursor:pointer">Скасувати</button>'
            '<button type="button" id="dirBrowserConfirm"'
            ' style="background:#4caf50;color:#fff;border:none;'
            'padding:8px 22px;border-radius:6px;cursor:pointer;font-weight:700">✅ Вибрати цю папку</button>'
            '</div></div></div>'
        )
        return mark_safe(html)

    upload_widget.short_description = '📤 Завантажити документи'
    
    def change_view(self, request, object_id, form_url='', extra_context=None):
        """Зберігаємо request для доступу в методах"""
        self.request = request
        extra_context = extra_context or {}
        try:
            obj = self.get_object(request, object_id)
            if obj and obj.source == "digikey" and obj.status in ("received", "processing"):
                from django.urls import reverse as _rev
                extra_context["digikey_confirm_url"] = _rev(
                    "admin:bots_digikeyconfig_confirm_order",
                    args=[obj.order_number],
                )
        except Exception:
            pass
        return super().change_view(request, object_id, form_url, extra_context)
    
    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
    
    def get_form(self, request, obj=None, **kwargs):
        """Підставляє збережений шлях із session або settings; динамічні джерела."""
        from django.conf import settings as _s
        form = super().get_form(request, obj, **kwargs)

        # Динамічний dropdown для поля source
        sources = list(SalesSource.objects.order_by("order", "name").values_list("slug", "name"))
        if sources:
            form.base_fields["source"].widget = forms.Select(
                choices=sources,
                attrs={"style": "max-width:300px"},
            )

        # Пріоритет: session → settings.LOCAL_DOCS_BASE_PATH → ''
        saved = request.session.get('local_base_path', '')
        default = getattr(_s, 'LOCAL_DOCS_BASE_PATH', '')
        form.base_fields['local_base_path'].initial = saved or default

        return form
    
    # ── Колонки ────────────────────────────────────────────────────────────────

    def deadline_display(self, obj):
        """Дедлайн відправки з підсвіткою та днями залишилось"""
        if not obj.shipping_deadline:
            return format_html('<span style="color:#7d8590">—</span>')
        
        # Якщо вже відправлено - показати тільки дату
        if obj.shipped_at:
            return format_html(
                '<div style="font-size:12px;color:var(--text-muted)">{}</div>',
                obj.shipping_deadline.strftime('%d.%m.%Y')
            )
        
        # Розрахувати дні до дедлайну
        today = date.today()
        delta = (obj.shipping_deadline - today).days
        
        # Кольори залежно від термінів
        if delta < 0:
            # Прострочено
            color = '#f85149'
            bg = 'rgba(248,81,73,0.1)'
            icon = '🔴'
            text = f'Прострочено {abs(delta)}д'
        elif delta == 0:
            # Сьогодні
            color = '#e65100'
            bg = 'rgba(230,81,0,0.12)'
            icon = '⚠️'
            text = 'Сьогодні!'
        elif delta <= 2:
            # Критично (1-2 дні)
            color = '#e65100'
            bg = 'rgba(230,81,0,0.1)'
            icon = '⏰'
            text = f'Залишилось {delta}д'
        elif delta <= 5:
            # Увага (3-5 днів)
            color = '#58a6ff'
            bg = 'rgba(88,166,255,0.08)'
            icon = '📅'
            text = f'Залишилось {delta}д'
        else:
            # Нормально (>5 днів)
            color = '#3fb950'
            bg = 'rgba(63,185,80,0.08)'
            icon = '✅'
            text = f'Залишилось {delta}д'
        
        return format_html(
            '<div style="font-size:12px;line-height:1.5">'
            '<div style="color:var(--text);font-weight:700;font-family:monospace">{}</div>'
            '<div style="margin-top:3px;padding:2px 8px;border-radius:4px;'
            'background:{};color:{};font-size:11px;font-weight:700;display:inline-block">'
            '{} {}</div></div>',
            obj.shipping_deadline.strftime('%d.%m.%Y'),
            bg,
            color,
            icon,
            text
        )

    deadline_display.short_description = '📦 Дедлайн'
    deadline_display.admin_order_field = 'shipping_deadline'


    def customer_link_display(self, obj):
        """Посилання на клієнта в CRM з RFM статусом."""
        from django.utils.html import format_html
        from django.urls import reverse
        
        if not obj.email and not obj.client:
            return "—"
        
        from crm.models import Customer
        from django.db.models import Q
        
        # Шукаємо клієнта
        customer = None
        if obj.customer_key:
            customer = Customer.objects.filter(external_key=obj.customer_key).first()
        if not customer and obj.email:
            customer = Customer.objects.filter(email=obj.email).first()
        if not customer and obj.client:
            customer = Customer.objects.filter(name__iexact=obj.client).first()
        
        if not customer:
            return format_html(
                '<div style="opacity:0.6">{}<br><small>⚠️ Не в CRM</small></div>',
                obj.client or obj.email
            )
        
        # Отримуємо RFM з обробкою помилок
        try:
            rfm = customer.rfm_score()
            segment = rfm.get('segment', '👤 Customer')
        except Exception as e:
            # Якщо rfm_score() падає - показуємо просто ім'я
            return format_html(
                '<a href="{}" style="text-decoration:none;color:#64b5f6">'
                '<div style="font-weight:600">{}</div>'
                '<div style="font-size:10px;color:#999">⚠️ RFM error</div></a>',
                reverse('admin:crm_customer_change', args=[customer.pk]),
                customer.company or customer.name
            )

        icons = {
            '🏆 Champions': '🏆', '💎 Loyal': '💎', '⭐ Potential': '⭐',
            '🔄 Regular': '🔄', '😴 At Risk': '😴', '💤 Hibernating': '💤', '🆕 New': '🆕',
        }
        colors = {
            '🏆 Champions': '#4caf50', '💎 Loyal': '#2196f3', '⭐ Potential': '#ff9800',
            '🔄 Regular': '#9c27b0', '😴 At Risk': '#f44336', '💤 Hibernating': '#757575', '🆕 New': '#00bcd4',
        }

        url = reverse('admin:crm_customer_change', args=[customer.pk])
        display_name = customer.company or customer.name
        sub_name = customer.name if customer.company else ""

        return format_html(
            '<a href="{}" style="text-decoration:none;color:#64b5f6">'
            '<div style="font-weight:600">{}</div>'
            '{}'
            '<div style="font-size:10px;color:{}">{}</div></a>',
            url,
            display_name,
            format_html('<div style="font-size:11px;color:#9aafbe">{}</div>', sub_name) if sub_name else "",
            colors.get(segment, '#616161'),
            segment,
        )
    customer_link_display.short_description = "Клієнт"

    def order_date_fmt(self, obj):
        if not obj.order_date:
            return format_html('<span style="color:#999">—</span>')
        return format_html('<span style="white-space:nowrap">{}</span>',
                           obj.order_date.strftime("%d.%m.%Y"))
    order_date_fmt.short_description = "Дата"
    order_date_fmt.admin_order_field = "order_date"

    def country_display(self, obj):
        from config.country_utils import display_country
        return display_country(obj.addr_country or "")
    country_display.short_description = "Країна"
    country_display.admin_order_field = "addr_country"

    def source_badge(self, obj):
        if not hasattr(self, '_source_map'):
            self._source_map = {
                s.slug: (s.color, s.name)
                for s in SalesSource.objects.all()
            }
        color, name = self._source_map.get(obj.source, ("#607d8b", obj.source))
        return format_html(
            '<span style="background:{};color:#fff;padding:2px 8px;border-radius:10px;'
            'font-size:11px;white-space:nowrap">{}</span>', color, name)
    source_badge.short_description = "Джерело"

    def status_badge(self, obj):
        colors = {"received": "#1565c0", "processing": "#e65100",
                  "shipped": "#2e7d32", "delivered": "#00695c", "cancelled": "#c62828"}
        color = colors.get(obj.status, "#757575")
        return format_html(
            '<span style="background:{};color:#fff;padding:3px 10px;border-radius:12px;'
            'font-size:11px;font-weight:bold;white-space:nowrap">{}</span>',
            color, obj.get_status_display())
    status_badge.short_description = "Статус"
    status_badge.admin_order_field = "status"

    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related('lines__product')

    def items_count(self, obj):
        return len(obj.lines.all())
    items_count.short_description = "Поз."

    def items_summary(self, obj):
        lines = sorted(obj.lines.all(), key=lambda l: l.pk)
        if not lines:
            return mark_safe('<span style="opacity:.4">—</span>')
        parts = []
        for line in lines[:4]:
            sku = line.sku_raw or (line.product.sku if line.product else '?')
            q = line.qty
            qty_str = str(int(q)) if q == int(q) else str(q)
            parts.append(
                f'<span style="font-weight:700">{sku}</span>'
                f'<span style="opacity:.6">&nbsp;×{qty_str}</span>'
            )
        if len(lines) > 4:
            parts.append(f'<span style="opacity:.45">+{len(lines) - 4} ще</span>')
        html = '<br>'.join(parts)
        return mark_safe(f'<span style="font-size:12px;font-family:monospace;line-height:1.7">{html}</span>')
    items_summary.short_description = "Товари"

    def order_total(self, obj):
        """Показує суму з валютою."""
        try:
            total = obj.lines.aggregate(t=Sum("total_price"))["t"]
        except Exception:
            total = None
        
        if not total:
            return "—"
        
        # Визначаємо символ валюти
        currency_symbols = {
            'USD': '$',
            'EUR': '€',
            'GBP': '£',
        }
        symbol = currency_symbols.get(obj.currency, obj.currency or '€')
        
        return format_html('<b>{}{}</b>', symbol, f"{float(total):.2f}")
    order_total.short_description = "Сума"

    def shipped_badge(self, obj):
        if obj.shipped_at:
            return format_html('<span style="color:#2e7d32;font-weight:700;white-space:nowrap">✅ {}</span>',
                               obj.shipped_at.strftime("%d.%m.%Y"))
        return format_html('<span style="color:#e65100;font-weight:700;white-space:nowrap">⏳ Очікує</span>')
    shipped_badge.short_description = "Відправлено"
    shipped_badge.admin_order_field = "shipped_at"

    def stock_warning(self, obj):
        if obj.shipped_at:
            return format_html('<span style="color:#999">—</span>')
        from inventory.models import InventoryTransaction
        problems, oks = [], 0
        for line in obj.lines.filter(product__isnull=False):
            result = InventoryTransaction.objects.filter(
                product=line.product).aggregate(total=Sum('qty'))
            stock = float(result['total'] or 0)
            needed = float(line.qty or 0)
            if stock <= 0:
                problems.append(f"🚫 {line.product.sku}: 0")
            elif stock < needed:
                problems.append(f"⚠️ {line.product.sku}: {int(stock)}/{int(needed)}")
            else:
                oks += 1
        if problems:
            tip = " | ".join(problems)
            label = problems[0] if len(problems) == 1 else f"{len(problems)} проблем"
            return format_html(
                '<span style="color:#f44336;font-size:11px;font-weight:bold" title="{}">'
                '🚫 {}</span>', tip, label)
        if oks:
            return format_html('<span style="color:#4caf50;font-size:11px">✅ OK</span>')
        return format_html('<span style="color:#999;font-size:11px">—</span>')
    stock_warning.short_description = "Склад"

    def customer_link(self, obj):
        if not obj.customer:
            return "—"
        from django.urls import reverse
        url = reverse("admin:crm_customer_change", args=[obj.customer.pk])
        return format_html('<a href="{}">👤 {}</a>', url, obj.customer.name)
    customer_link.short_description = "Клієнт CRM"

    # ── Кнопки етикеток у списку ───────────────────────────────────────────────

    def label_buttons_list(self, obj):
        from pathlib import Path
        from django.conf import settings
        labels_dir = Path(getattr(settings, 'LABELS_DIR', Path(settings.BASE_DIR) / 'labels'))
        buttons = []
        for line in obj.lines.all():
            sku = line.product.sku if line.product else line.sku_raw
            qty = int(line.qty or 1)
            if not sku:
                continue
            # Нормалізація: прибрати пробіли + замінити Unicode-дефіси на ASCII
            sku = sku.strip().replace('\u2013', '-').replace('\u2014', '-').replace('\u2010', '-').replace('\u2011', '-')
            found = False
            sku_up = sku.upper()
            for f in labels_dir.glob('*.dymo'):
                s = f.stem.upper()
                if s == sku_up or (s.startswith(sku_up) and len(s) > len(sku_up) and s[len(sku_up)] in (' ', '_')):
                    found = True
                    break
            if found:
                url = f"/labels/serve/{sku}/?qty={qty}"
                buttons.append(
                    f'<a href="{url}" '
                    f'style="display:inline-block;margin:2px;background:#1976d2;color:#fff;'
                    f'padding:3px 8px;border-radius:6px;font-size:11px;text-decoration:none;'
                    f'white-space:nowrap">🖨️ {sku}</a>')
            else:
                buttons.append(
                    f'<span style="display:inline-block;margin:2px;background:#bdbdbd;color:#fff;'
                    f'padding:3px 8px;border-radius:6px;font-size:11px;white-space:nowrap" '
                    f'title="Пошук: {labels_dir} | SKU={repr(sku)}">❌ {sku}</span>')
        return mark_safe("".join(buttons)) if buttons else format_html('<span style="color:#999">—</span>')
    label_buttons_list.short_description = "🏷️ Етикетки"

    # ── Детальний блок етикеток ────────────────────────────────────────────────

    def label_buttons_detail(self, obj):
        from pathlib import Path
        from django.conf import settings
        labels_dir = Path(getattr(settings, 'LABELS_DIR', Path(settings.BASE_DIR) / 'labels'))
        rows = []
        for line in obj.lines.all():
            sku = line.product.sku if line.product else line.sku_raw
            name = line.product.name if line.product else (line.sku_raw or '—')
            qty = int(line.qty or 1)
            if not sku:
                continue
            sku = sku.strip().replace('\u2013', '-').replace('\u2014', '-').replace('\u2010', '-').replace('\u2011', '-')
            label_path = None
            sku_up = sku.upper()
            for f in labels_dir.glob('*.dymo'):
                s = f.stem.upper()
                if s == sku_up:
                    label_path = f
                    break
                if label_path is None and s.startswith(sku_up) and len(s) > len(sku_up) and s[len(sku_up)] in (' ', '_'):
                    label_path = f
            if label_path:
                url = f"/labels/serve/{sku}/?qty={qty}"
                btn = (f'<a href="{url}" target="_blank" '
                       f'style="background:#1976d2;color:#fff;padding:8px 16px;'
                       f'border-radius:8px;text-decoration:none;font-weight:bold">'
                       f'🖨️ Друкувати ({qty} шт.)</a>')
                st = f'<span style="color:#4caf50;margin-left:10px">✅ {label_path.name}</span>'
            else:
                btn = '<span style="color:rgba(150,150,150,0.9);font-style:italic">Файл не завантажено</span>'
                st = (f'<span style="color:#f44336;margin-left:10px">'
                      f'❌ {sku}.dymo — завантажте нижче</span>')
            rows.append(
                f'<tr style="border-bottom:1px solid rgba(128,128,128,0.15)">'
                f'<td style="padding:10px;font-weight:bold">{sku}</td>'
                f'<td style="padding:10px;opacity:0.85">{name}</td>'
                f'<td style="padding:10px;text-align:center;font-weight:bold">{qty}</td>'
                f'<td style="padding:10px">{btn}{st}</td></tr>')
        if not rows:
            return mark_safe('<p style="color:#999">Немає позицій</p>')
        return mark_safe(
            '<table style="border-collapse:collapse;width:100%;'
            'border-radius:8px;overflow:hidden;border:1px solid rgba(128,128,128,0.2)">'
            '<thead><tr style="background:rgba(21,101,192,0.85);color:#e3f2fd">'
            '<th style="padding:10px;text-align:left">SKU</th>'
            '<th style="padding:10px;text-align:left">Назва</th>'
            '<th style="padding:10px;text-align:center">К-сть</th>'
            '<th style="padding:10px;text-align:left">Дія</th>'
            '</tr></thead><tbody>' + "".join(rows) + '</tbody></table>')
    label_buttons_detail.short_description = "Друк етикеток"

    # ── Завантаження ───────────────────────────────────────────────────────────

    def label_upload_widget(self, obj):
        # CSRF через hidden input Django, не через cookie
        return mark_safe('''
        <div style="background:rgba(255,255,255,0.05);border:1px solid rgba(255,255,255,0.1);
                    padding:16px;border-radius:8px">
          <p style="margin:0 0 10px;font-weight:bold;opacity:0.9">
            📤 Завантажити або оновити .dymo файли</p>
          <input type="file" id="dymoFiles" multiple accept=".dymo"
                 style="margin-bottom:10px;display:block;opacity:0.85">
          <button type="button" id="dymoUploadBtn"
                  style="background:#4caf50;color:#fff;border:none;padding:8px 20px;
                         border-radius:6px;cursor:pointer;font-size:14px;font-weight:bold">
            ⬆️ Завантажити на сервер
          </button>
          <div id="dymoResult" style="margin-top:10px;font-size:13px"></div>
        </div>
        <script>
        document.getElementById("dymoUploadBtn").addEventListener("click", function() {
            var input = document.getElementById("dymoFiles");
            var result = document.getElementById("dymoResult");
            if (!input.files.length) {
                result.innerHTML = "<span style=\'color:#f44336\'>Оберіть файли!</span>";
                return;
            }
            var formData = new FormData();
            for (var i = 0; i < input.files.length; i++) {
                formData.append("labels", input.files[i]);
            }
            // CSRF з Django hidden input
            var csrfEl = document.querySelector("[name=csrfmiddlewaretoken]");
            var csrf = csrfEl ? csrfEl.value : "";
            result.innerHTML = "⏳ Завантаження...";
            fetch("/labels/upload/", {
                method: "POST",
                headers: {"X-CSRFToken": csrf},
                body: formData
            })
            .then(function(r) { return r.json(); })
            .then(function(data) {
                var html = "<ul style=\'margin:5px 0;padding-left:20px\'>";
                data.results.forEach(function(r) {
                    var icon = r.status === "created" ? "✅" : r.status === "updated" ? "🔄" : "❌";
                    var color = r.status === "error" ? "#f44336" : "#4caf50";
                    html += "<li style=\'color:" + color + "\'>" + icon + " " + r.name + ": " + r.status;
                    if (r.size) html += " (" + (r.size/1024).toFixed(1) + " KB)";
                    html += "</li>";
                });
                html += "</ul><button onclick=\'location.reload()\' "
                     + "style=\'background:#1976d2;color:#fff;border:none;padding:6px 14px;"
                     + "border-radius:4px;cursor:pointer;margin-top:6px\'>🔄 Оновити</button>";
                result.innerHTML = html;
            })
            .catch(function(e) {
                result.innerHTML = "<span style=\'color:#f44336\'>Помилка: " + e + "</span>";
            });
        });
        </script>
        ''')
    label_upload_widget.short_description = "Завантажити етикетки"

    # ── Склад ──────────────────────────────────────────────────────────────────

    def stock_summary(self, obj):
        from inventory.models import InventoryTransaction
        lines = obj.lines.all()
        if not lines:
            return "Немає позицій"
        rows, has_problems = [], False
        for line in lines:
            sku = line.product.sku if line.product else line.sku_raw
            name = line.product.name if line.product else "— SKU не знайдено —"
            needed = float(line.qty or 0)
            if line.product:
                result = InventoryTransaction.objects.filter(
                    product=line.product).aggregate(total=Sum('qty'))
                stock = float(result['total'] or 0)
                if stock <= 0:
                    s = "🚫 НЕМАЄ"
                    row_style = "border-left:4px solid #f44336"
                    tc = "#f44336"
                    has_problems = True
                elif stock < needed:
                    s = f"⚠️ {int(stock)} / треба {int(needed)}"
                    row_style = "border-left:4px solid #ff9800"
                    tc = "#ff9800"
                    has_problems = True
                else:
                    s = f"✅ {int(stock)} шт."
                    row_style = "border-left:4px solid #4caf50"
                    tc = "#4caf50"
            else:
                s = "⚠️ Не в базі"
                row_style = "border-left:4px solid #ff9800"
                tc = "#ff9800"

            rows.append(
                f"<tr style='{row_style}'>"
                f"<td style='padding:10px;font-weight:bold'>{sku}</td>"
                f"<td style='padding:10px'>{name}</td>"
                f"<td style='padding:10px;text-align:center'>{int(needed)}</td>"
                f"<td style='padding:10px;color:{tc};font-weight:bold'>{s}</td></tr>"
            )

        summary_icon = "⚠️" if has_problems else "✅"
        summary_text = "Є проблеми з наявністю!" if has_problems else "Всі товари є на складі"
        summary_color = "#f44336" if has_problems else "#4caf50"

        return mark_safe(
            f'<div style="border-left:4px solid {summary_color};padding:10px 16px;'
            f'margin-bottom:12px;border-radius:4px;font-weight:bold;'
            f'color:{summary_color}">'
            f'{summary_icon} {summary_text}</div>'
            '<table style="border-collapse:collapse;width:100%;'
            'border-radius:8px;overflow:hidden;border:1px solid rgba(128,128,128,0.2)">'
            '<thead><tr style="background:rgba(55,71,79,0.9);color:#eceff1">'
            '<th style="padding:10px;text-align:left">SKU</th>'
            '<th style="padding:10px;text-align:left">Назва</th>'
            '<th style="padding:10px;text-align:center">Потрібно</th>'
            '<th style="padding:10px;text-align:left">Склад</th>'
            '</tr></thead>'
            '<tbody style="background:transparent">'
            + "".join(rows) +
            '</tbody></table>'
        )
    stock_summary.short_description = "📦 Залишки на складі"
    
    # ══════════════════════════════════════════════════════════════════════════
    # MANUAL IMPORT EXCEL
    # ══════════════════════════════════════════════════════════════════════════
    
    change_list_template = 'admin/sales/salesorder_changelist.html'
    change_form_template = 'admin/sales/salesorder_change_form.html'

    def get_changeform_initial_data(self, request):
        initial = super().get_changeform_initial_data(request)
        mapping = {
            '_prefill_client':          'client',
            '_prefill_contact_name':    'contact_name',
            '_prefill_email':           'email',
            '_prefill_phone':           'phone',
            '_prefill_addr_street':     'addr_street',
            '_prefill_addr_city':       'addr_city',
            '_prefill_addr_zip':        'addr_zip',
            '_prefill_addr_state':      'addr_state',
            '_prefill_addr_country':    'addr_country',
            '_prefill_shipping_address':'shipping_address',
        }
        for param, field in mapping.items():
            val = request.GET.get(param)
            if val:
                initial[field] = val
        return initial

    def get_urls(self):
        """Додаємо URL для manual import, завантаження документів та PDF-генерації."""
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/',
                 self.admin_site.admin_view(self.import_excel_view),
                 name='sales_salesorder_import'),
            path('<int:pk>/upload-docs/',
                 self.admin_site.admin_view(self.upload_docs_view),
                 name='sales_salesorder_upload_docs'),
            path('browse-dir/',
                 self.admin_site.admin_view(self.browse_dir_view),
                 name='sales_salesorder_browse_dir'),
            # PDF document generators (download + ?action=save)
            path('<int:pk>/doc/packing-list/',
                 self.admin_site.admin_view(self._doc_packing_list),
                 name='sales_salesorder_doc_packing_list'),
            path('<int:pk>/doc/proforma/',
                 self.admin_site.admin_view(self._doc_proforma),
                 name='sales_salesorder_doc_proforma'),
            path('<int:pk>/doc/customs/',
                 self.admin_site.admin_view(self._doc_customs),
                 name='sales_salesorder_doc_customs'),
            # Delete uploaded document
            path('<int:pk>/doc/delete/',
                 self.admin_site.admin_view(self.delete_doc_view),
                 name='sales_salesorder_delete_doc'),
            # AJAX: refresh documents panel
            path('<int:pk>/doc/list/',
                 self.admin_site.admin_view(self.doc_list_view),
                 name='sales_salesorder_doc_list'),
        ]
        return custom_urls + urls

    # ── PDF document views ─────────────────────────────────────────────────────

    def _doc_packing_list(self, request, pk):
        order = get_object_or_404(SalesOrder, pk=pk)
        from sales.doc_generators import generate_packing_list
        overrides = {k: v for k, v in request.GET.items() if k != 'action'}
        buf = generate_packing_list(order, overrides=overrides or None)
        if request.GET.get('action') == 'save':
            return self._save_pdf_to_media(order, buf, f'PackingList-{order.order_number}.pdf')
        response = HttpResponse(buf.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="PackingList-{order.order_number}.pdf"'
        )
        return response

    def _doc_proforma(self, request, pk):
        order = get_object_or_404(SalesOrder, pk=pk)
        from sales.doc_generators import generate_proforma
        overrides = {k: v for k, v in request.GET.items() if k != 'action'}
        buf = generate_proforma(order, overrides=overrides or None)
        if request.GET.get('action') == 'save':
            return self._save_pdf_to_media(order, buf, f'ProformaInvoice-{order.order_number}.pdf')
        response = HttpResponse(buf.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="ProformaInvoice-{order.order_number}.pdf"'
        )
        return response

    def _doc_customs(self, request, pk):
        order = get_object_or_404(SalesOrder, pk=pk)
        from sales.doc_generators import generate_customs
        overrides = {k: v for k, v in request.GET.items() if k != 'action'}
        buf = generate_customs(order, overrides=overrides or None)
        if request.GET.get('action') == 'save':
            return self._save_pdf_to_media(order, buf, f'CustomsDeclaration-{order.order_number}.pdf')
        response = HttpResponse(buf.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="CustomsDeclaration-{order.order_number}.pdf"'
        )
        return response

    def _save_pdf_to_media(self, order, buf, filename):
        """Зберігає PDF в media/orders/{source}/{order_number}/ та повертає JSON."""
        from django.http import JsonResponse
        from django.conf import settings
        dest_dir = settings.MEDIA_ROOT / 'orders' / (order.source or 'manual') / order.order_number
        dest_dir.mkdir(parents=True, exist_ok=True)
        (dest_dir / filename).write_bytes(buf.getvalue())
        rel_url = (
            f'{settings.MEDIA_URL}orders/{order.source or "manual"}'
            f'/{order.order_number}/{filename}'
        )
        return JsonResponse({'ok': True, 'filename': filename, 'url': rel_url})

    def delete_doc_view(self, request, pk):
        """AJAX POST — видаляє один файл із media/orders/{source}/{order_number}/."""
        from django.http import JsonResponse
        from django.conf import settings
        if request.method != 'POST':
            return JsonResponse({'error': 'Method not allowed'}, status=405)
        order = get_object_or_404(SalesOrder, pk=pk)
        filename = request.POST.get('filename', '').strip()
        # Захист від path traversal
        if not filename or '/' in filename or '\\' in filename or '..' in filename:
            return JsonResponse({'error': 'Invalid filename'}, status=400)
        target = settings.MEDIA_ROOT / 'orders' / order.source / order.order_number / filename
        if not target.is_file():
            return JsonResponse({'error': 'File not found'}, status=404)
        target.unlink()
        return JsonResponse({'ok': True, 'deleted': filename})

    def doc_list_view(self, request, pk):
        """AJAX GET — повертає HTML-фрагмент панелі документів для оновлення без перезавантаження."""
        order = get_object_or_404(SalesOrder, pk=pk)
        return HttpResponse(self._docs_panel_html(order))

    def upload_docs_view(self, request, pk):
        """AJAX-ендпоінт: зберігає документи замовлення на сервер (і локально)."""
        from django.http import JsonResponse
        from django.conf import settings
        from datetime import date as _date

        if request.method != 'POST':
            return JsonResponse({'error': 'Method not allowed'}, status=405)

        try:
            order = SalesOrder.objects.get(pk=pk)
        except SalesOrder.DoesNotExist:
            return JsonResponse({'error': 'Замовлення не знайдено'}, status=404)

        files = request.FILES.getlist('documents')
        if not files:
            return JsonResponse({'error': 'Файли не вибрані'}, status=400)

        import os
        from pathlib import Path as _Path

        raw_path = request.POST.get('local_base_path', '').strip().strip('"\'')
        if not raw_path:
            raw_path = getattr(settings, 'LOCAL_DOCS_BASE_PATH', '')
        local_base_path = raw_path.strip().rstrip('/\\')

        if local_base_path:
            request.session['local_base_path'] = local_base_path

        server_base = settings.MEDIA_ROOT / 'orders' / order.source / order.order_number
        server_base.mkdir(parents=True, exist_ok=True)

        # Локальний шлях (server-side copy — Firefox fallback)
        local_dir = None
        local_error = None
        local_dir_str = ''
        if local_base_path:
            try:
                import os as _os
                from datetime import date as _date2
                normalized  = local_base_path.replace('\\', _os.sep).replace('/', _os.sep)
                path_obj    = _Path(normalized)
                source_slug = (order.source or 'manual').lower().replace(' ', '_')
                date_str    = _date2.today().strftime('%Y-%m-%d')

                if order.order_number and path_obj.name == order.order_number:
                    local_dir = path_obj
                elif path_obj.name.lower() == source_slug:
                    local_dir = path_obj / date_str / order.order_number
                else:
                    local_dir = path_obj / source_slug / date_str / order.order_number
                local_dir.mkdir(parents=True, exist_ok=True)
                local_dir_str = str(local_dir)
            except Exception as e:
                local_error = str(e)
                local_dir = None

        results = []
        for f in files:
            try:
                dest = server_base / f.name
                with dest.open('wb+') as fh:
                    for chunk in f.chunks():
                        fh.write(chunk)

                local_status = None
                if local_base_path:
                    if local_error:
                        local_status = {'ok': False, 'msg': local_error}
                    elif local_dir:
                        try:
                            local_file = local_dir / f.name
                            shutil.copy2(str(dest), str(local_file))
                            local_status = {'ok': True, 'msg': str(local_file)}
                        except Exception as e:
                            local_status = {'ok': False, 'msg': str(e)}

                results.append({'name': f.name, 'status': 'saved', 'size': f.size, 'local': local_status})
            except Exception as e:
                results.append({'name': f.name, 'status': 'error', 'size': 0, 'error': str(e)})

        debug = {
            'local_base_path_received': local_base_path or '(порожньо)',
            'local_path_attempted': local_dir_str or '(не будувався)',
            'local_mkdir_error': local_error or None,
        }
        return JsonResponse({'results': results, 'debug': debug})

    def browse_dir_view(self, request):
        """AJAX — повертає список підпапок для серверного браузера директорій."""
        from django.http import JsonResponse
        import os

        path = request.GET.get('path', '').strip()
        if not path:
            path = request.session.get('local_base_path', '') or os.path.expanduser('~')

        path = os.path.normpath(path)
        if not os.path.isdir(path):
            parent = os.path.dirname(path)
            path = parent if os.path.isdir(parent) else os.path.expanduser('~')

        parent = os.path.dirname(path)
        if parent == path:
            parent = None  # корінь файлової системи

        dirs = []
        try:
            with os.scandir(path) as it:
                for entry in sorted(it, key=lambda e: e.name.lower()):
                    try:
                        if entry.is_dir(follow_symlinks=False):
                            dirs.append({'name': entry.name, 'path': entry.path})
                    except OSError:
                        pass
        except OSError:
            pass

        return JsonResponse({'current': path, 'parent': parent, 'dirs': dirs})

    def import_excel_view(self, request):
        """3-кроковий wizard імпорту замовлень з Excel з динамічним маппінгом колонок."""
        from collections import defaultdict
        from datetime import datetime

        # ── Поля БД для маппінгу ──────────────────────────────────────────────
        ORDER_FIELDS = [
            ("order_number",      "Номер замовлення * (ключ)"),
            ("source",            "Джерело (digikey/nova_post/manual)"),
            ("order_date",        "Дата замовлення"),
            ("shipped_at",        "Дата відправки"),
            ("shipping_deadline", "Дедлайн відправки"),
            ("shipping_courier",  "Кур'єр"),
            ("tracking_number",   "Трекінг номер"),
            ("lieferschein_nr",   "Lieferschein Nr"),
            ("shipping_region",   "Регіон доставки"),
            ("shipping_address",  "Адреса доставки (legacy)"),
            ("addr_street",       "Вулиця, будинок"),
            ("addr_city",         "Місто"),
            ("addr_zip",          "Поштовий індекс"),
            ("addr_state",        "Штат / провінція (ISO-2: CA/NY/TX)"),
            ("addr_country",      "Країна (ISO-2: DE/UA/PL)"),
            ("client",            "Клієнт / Компанія"),
            ("contact_name",      "Контактна особа"),
            ("phone",             "Телефон"),
            ("email",             "Email"),
            ("currency",          "Валюта замовлення"),
            ("shipping_cost",     "Вартість доставки"),
            ("shipping_currency", "Валюта доставки"),
            ("status",            "Статус"),
            ("document_type",     "Тип документу"),
        ]
        LINE_FIELDS = [
            ("sku_raw",         "SKU товару (рядок)"),
            ("qty",             "Кількість (рядок)"),
            ("unit_price",      "Ціна за одиницю"),
            ("total_price_line","Загальна сума рядка"),
            ("currency_line",   "Валюта рядка"),
        ]
        ALL_DB_FIELDS = ORDER_FIELDS + LINE_FIELDS
        ORDER_FIELD_KEYS = {k for k, _ in ORDER_FIELDS}
        LINE_FIELD_KEYS  = {k for k, _ in LINE_FIELDS}

        def parse_price(value):
            if not value:
                return None, None
            s = str(value).strip()
            su = s.upper()
            currency = 'USD' if ('$' in s or 'USD' in su) else ('EUR' if ('€' in s or 'EUR' in su) else None)
            price_str = re.sub(r'[^\d\.,]', '', s)
            if not price_str:
                return None, currency
            # Determine decimal separator: if both comma and dot present,
            # the last one is the decimal separator
            has_comma = ',' in price_str
            has_dot   = '.' in price_str
            if has_comma and has_dot:
                if price_str.rfind(',') > price_str.rfind('.'):
                    # European: 1.234,56 → remove dot thousands, comma→dot
                    price_str = price_str.replace('.', '').replace(',', '.')
                else:
                    # US: 1,234.56 → remove comma thousands
                    price_str = price_str.replace(',', '')
            else:
                # Only comma (European decimal) or only dot — normalise comma
                price_str = price_str.replace(',', '.')
            try:
                return (Decimal(price_str) if price_str else None), currency
            except InvalidOperation:
                return None, currency

        def convert_date(value):
            if not value:
                return None
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            if isinstance(value, str):
                v = value.strip()
                for fmt in ('%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y'):
                    try:
                        return datetime.strptime(v, fmt).date()
                    except ValueError:
                        continue
            return None

        ctx_base = dict(
            self.admin_site.each_context(request),
            title="📥 Імпорт замовлень з Excel",
            opts=self.model._meta,
            order_fields=ORDER_FIELDS,
            line_fields=LINE_FIELDS,
        )

        # ── STEP 1: форма завантаження ─────────────────────────────────────────
        if request.method == "GET" or request.POST.get("step") not in ("1", "2"):
            form = SalesExcelUploadForm()
            return render(request, "admin/sales/import_excel.html",
                          {**ctx_base, "step": 1, "form": form})

        # ── STEP 1 POST: аналіз файлу ─────────────────────────────────────────
        if request.POST.get("step") == "1":
            form = SalesExcelUploadForm(request.POST, request.FILES)
            if not form.is_valid():
                return render(request, "admin/sales/import_excel.html",
                              {**ctx_base, "step": 1, "form": form})

            uploaded = request.FILES["excel_file"]
            if not uploaded.name.lower().endswith(".xlsx"):
                form.add_error("excel_file", "Підтримується лише формат .xlsx")
                return render(request, "admin/sales/import_excel.html",
                              {**ctx_base, "step": 1, "form": form})

            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            for chunk in uploaded.chunks():
                tmp.write(chunk)
            tmp.close()
            request.session["sales_excel_import_tmp"] = tmp.name

            try:
                wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
            except Exception as e:
                os.unlink(tmp.name)
                form.add_error("excel_file", f"Не вдалося відкрити файл: {e}")
                return render(request, "admin/sales/import_excel.html",
                              {**ctx_base, "step": 1, "form": form})

            sheet_data = {}
            for sname in wb.sheetnames:
                ws = wb[sname]
                header_cols = []
                for row in ws.iter_rows(max_row=10, values_only=True):
                    non_empty = [c for c in row if c is not None and str(c).strip()]
                    if non_empty:
                        header_cols = [str(c).strip() if c is not None else "" for c in row]
                        break
                sheet_data[sname] = header_cols
            wb.close()

            return render(request, "admin/sales/import_excel.html", {
                **ctx_base,
                "step": 2,
                "sheet_names": wb.sheetnames,
                "sheet_data_json": json.dumps(sheet_data, ensure_ascii=False),
                "first_sheet": wb.sheetnames[0] if wb.sheetnames else "",
                "first_cols": sheet_data.get(wb.sheetnames[0], []) if wb.sheetnames else [],
                "file_name": uploaded.name,
            })

        # ── STEP 2 POST: виконання імпорту ────────────────────────────────────
        if request.POST.get("step") == "2":
            tmp_path = request.session.get("sales_excel_import_tmp")
            if not tmp_path or not os.path.exists(tmp_path):
                from django.contrib import messages as _msg
                _msg.error(request, "Сесія застаріла — завантажте файл знову.")
                return redirect(reverse("admin:sales_salesorder_import"))

            sheet_name    = request.POST.get("sheet_name", "")
            conflict_mode = request.POST.get("conflict_mode", "skip")
            dry_run       = bool(request.POST.get("dry_run"))
            default_source = request.POST.get("default_source", "manual").strip() or "manual"

            # Збираємо маппінг: col_idx → field_name
            mappings = {}
            for key, val in request.POST.items():
                if key.startswith("col_") and val and val != "--":
                    try:
                        mappings[int(key[4:])] = val
                    except ValueError:
                        pass

            if not any(v == "order_number" for v in mappings.values()):
                from django.contrib import messages as _msg
                _msg.error(request, "Необхідно вибрати колонку для поля 'Номер замовлення'.")
                return redirect(reverse("admin:sales_salesorder_import"))

            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]

            stats = {"created": 0, "updated": 0, "skipped": 0,
                     "lines_created": 0, "errors": []}

            # Групуємо рядки по order_number
            orders_data = defaultdict(lambda: {"header": {}, "lines": []})
            header_skipped = False

            for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                if not header_skipped:
                    header_skipped = True
                    continue

                row_vals = {}
                for idx, field in mappings.items():
                    row_vals[field] = row[idx] if idx < len(row) else None

                order_num_raw = row_vals.get("order_number")
                if not order_num_raw or str(order_num_raw).strip() == "":
                    continue
                order_key = str(order_num_raw).strip()

                entry = orders_data[order_key]

                # Header — заповнюємо тільки перший раз
                if not entry["header"]:
                    hdr = {"order_number": order_key,
                           "source": default_source}
                    for field in ORDER_FIELD_KEYS - {"order_number"}:
                        raw = row_vals.get(field)
                        if raw is None:
                            continue
                        if field in ("order_date", "shipped_at", "shipping_deadline"):
                            hdr[field] = convert_date(raw)
                        elif field == "shipping_cost":
                            price, curr = parse_price(raw)
                            hdr["shipping_cost"] = price or Decimal("0")
                            if "shipping_currency" not in row_vals or not row_vals.get("shipping_currency"):
                                hdr["shipping_currency"] = curr or "EUR"
                        elif field == "source":
                            hdr["source"] = str(raw).strip() or default_source
                        elif field == "shipping_courier":
                            from sales.utils import normalize_courier
                            hdr["shipping_courier"] = normalize_courier(str(raw).strip())
                        else:
                            hdr[field] = str(raw).strip() if raw is not None else ""
                    entry["header"] = hdr

                # Line item
                sku_raw = row_vals.get("sku_raw")
                qty_raw = row_vals.get("qty")
                if sku_raw and str(sku_raw).strip():
                    try:
                        qty = Decimal(str(qty_raw).replace(",", ".")) if qty_raw else Decimal("1")
                    except InvalidOperation:
                        qty = Decimal("1")
                    unit_price, u_curr = parse_price(row_vals.get("unit_price"))
                    total_price, t_curr = parse_price(row_vals.get("total_price_line"))
                    currency_raw = row_vals.get("currency_line")
                    line_curr = (str(currency_raw).strip() if currency_raw else None) or u_curr or t_curr or "USD"
                    entry["lines"].append({
                        "sku": str(sku_raw).strip(),
                        "qty": qty,
                        "unit_price": unit_price,
                        "total_price": total_price,
                        "currency": line_curr,
                    })
            wb.close()

            # ── Зберігаємо в БД ───────────────────────────────────────────────
            from inventory.models import Product, ProductAlias
            from crm.models import Customer

            with transaction.atomic():
                for order_key, data in orders_data.items():
                    try:
                        hdr = data["header"]
                        existing = SalesOrder.objects.filter(
                            order_number=order_key).first()

                        if existing and conflict_mode == "skip":
                            stats["skipped"] += 1
                            continue

                        # Customer
                        email = hdr.get("email", "")
                        client = hdr.get("client", "")
                        if email or client:
                            cust_key = Customer.generate_key(
                                email or client, client or email)
                            hdr["customer_key"] = cust_key
                            Customer.objects.get_or_create(
                                external_key=cust_key,
                                defaults={
                                    "name": client or (email.split("@")[0] if email else ""),
                                    "email": email,
                                    "phone": hdr.get("phone", ""),
                                    "country": (hdr.get("addr_country") or hdr.get("shipping_region") or "")[:2],
                                    "source": hdr.get("source", default_source),
                                },
                            )

                        if existing:
                            for k, v in hdr.items():
                                setattr(existing, k, v)
                            existing.save()
                            existing.lines.all().delete()
                            order = existing
                            stats["updated"] += 1
                        else:
                            order = SalesOrder.objects.create(**hdr)
                            stats["created"] += 1

                        for line in data["lines"]:
                            product = (
                                Product.objects.filter(sku=line["sku"]).first()
                                or ProductAlias.objects.filter(alias=line["sku"])
                                          .select_related("product")
                                          .first()
                                and ProductAlias.objects.filter(alias=line["sku"])
                                          .select_related("product")
                                          .first().product
                            )
                            SalesOrderLine.objects.create(
                                order=order,
                                product=product,
                                sku_raw=line["sku"],
                                qty=line["qty"],
                                unit_price=line["unit_price"],
                                total_price=line["total_price"],
                                currency=line["currency"],
                            )
                            stats["lines_created"] += 1

                    except Exception as e:
                        stats["errors"].append(f"Замовлення {order_key}: {e}")

                if dry_run:
                    transaction.set_rollback(True)

            return render(request, "admin/sales/import_excel.html", {
                **ctx_base,
                "step": 3,
                "stats": stats,
                "dry_run": dry_run,
            })

        # Fallback
        return redirect(reverse("admin:sales_salesorder_import"))


# ── Inject sales stats into sales app_index context ────────────────────────────

def _get_sales_stats():
    try:
        from django.db.models import Count
        today = date.today()
        import calendar
        month_start = today.replace(day=1)

        qs = SalesOrder.objects.values("status").annotate(n=Count("pk"))
        by_status = {row["status"]: row["n"] for row in qs}

        overdue = SalesOrder.objects.filter(
            shipping_deadline__lt=today,
            status__in=["received", "processing"],
        ).count()

        shipped_month = SalesOrder.objects.filter(
            status="shipped",
            shipped_at__gte=month_start,
        ).count()

        return {
            "received":      by_status.get("received", 0),
            "processing":    by_status.get("processing", 0),
            "overdue":       overdue,
            "shipped_month": shipped_month,
        }
    except Exception:
        return {"received": "—", "processing": "—", "overdue": "—", "shipped_month": "—"}


_orig_sales_app_index = admin.site.app_index


def _sales_app_index(request, app_label, extra_context=None):
    if app_label == "sales":
        extra_context = extra_context or {}
        extra_context["sales_stats"] = _get_sales_stats()
    return _orig_sales_app_index(request, app_label, extra_context)


admin.site.app_index = _sales_app_index
