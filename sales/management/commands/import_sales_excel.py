"""
python manage.py import_sales_excel Post_FIXED_v5.xlsx

Імпортує дані з Excel файлу в Sales Orders.
Підтримує 3 листи: digikey, NovaPost, Other
"""
import sys
from decimal import Decimal
from datetime import datetime
from django.core.management.base import BaseCommand
from django.utils import timezone
from django.db import transaction
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Імпортувати продажі з Excel файлу"

    def add_arguments(self, parser):
        parser.add_argument('file', type=str, help='Шлях до Excel файлу')
        parser.add_argument(
            '--sheet',
            type=str,
            help='Назва листа (digikey/NovaPost/Other). Без параметра — всі листи'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Тільки показати дані без збереження'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        sheet_name = options.get('sheet')
        dry_run = options['dry_run']

        try:
            wb = load_workbook(file_path, data_only=True)
        except FileNotFoundError:
            self.stderr.write(self.style.ERROR(f'❌ Файл не знайдено: {file_path}'))
            return

        sheets = [sheet_name] if sheet_name else wb.sheetnames
        
        total_created = 0
        total_skipped = 0

        for sheet in sheets:
            if sheet not in wb.sheetnames:
                self.stderr.write(self.style.WARNING(f'⚠️ Лист "{sheet}" не знайдено'))
                continue

            self.stdout.write(f'\n📄 Обробка листа: {sheet}')
            created, skipped = self.import_sheet(wb[sheet], dry_run)
            total_created += created
            total_skipped += skipped

        if dry_run:
            self.stdout.write(self.style.WARNING(
                f'\n[DRY RUN] Нічого не збережено в БД'
            ))
        else:
            self.stdout.write(self.style.SUCCESS(
                f'\n✅ Готово!\n'
                f'   Створено замовлень: {total_created}\n'
                f'   Пропущено (дублікати): {total_skipped}'
            ))

    def import_sheet(self, ws, dry_run=False):
        """Імпортує один лист Excel."""
        from sales.models import SalesOrder, SalesOrderLine
        from inventory.models import Product, ProductAlias

        # Mapping колонок (базується на аналізі)
        COL = {
            'order_date': 1,          # A - Order Date
            'shipped_at': 2,          # B - Shipping
            'courier': 3,             # C - Shipping Courier
            'tracking': 4,            # D - tracking number
            'lieferschein': 5,        # E - Lieferschein-Nr
            'region': 6,              # F - Shipping Region
            'client': 7,              # G - client
            'address': 8,             # H - Shipping Address
            'product_number': 9,      # I - product namber
            'product_name': 10,       # J - product
            'qty': 11,                # K - QTY
            'unit_price': 12,         # L - Unit Price
            'total_value': 13,        # M - Total Value
            'shipping_cost': 14,      # N - Packing and shipping costs
            'source': 15,             # O - order via
            'order_number': 16,       # P - Sales Order
            'email': 17,              # Q - Email
            'deadline': 18,           # R - Shipping Deadline
            'firma': 19,              # S - Firma
            'name': 20,               # T - Name
            'phone': 21,              # U - Phone
        }

        created = 0
        skipped = 0

        # Групуємо рядки по Sales Order номеру
        orders_data = {}
        
        for row_idx in range(2, ws.max_row + 1):
            row = [ws.cell(row_idx, c).value for c in range(1, 22)]
            
            order_number = self._clean(row[COL['order_number'] - 1])
            if not order_number:
                continue  # пропустити рядки без номера замовлення

            if order_number not in orders_data:
                orders_data[order_number] = {
                    'order_date': row[COL['order_date'] - 1],
                    'shipped_at': row[COL['shipped_at'] - 1],
                    'courier': self._clean(row[COL['courier'] - 1]),
                    'tracking': self._clean(row[COL['tracking'] - 1]),
                    'lieferschein': self._clean(row[COL['lieferschein'] - 1]),
                    'region': self._clean(row[COL['region'] - 1]),
                    'client': self._clean(row[COL['client'] - 1]),
                    'address': self._clean(row[COL['address'] - 1]),
                    'email': self._clean(row[COL['email'] - 1]),
                    'deadline': row[COL['deadline'] - 1],
                    'firma': self._clean(row[COL['firma'] - 1]),
                    'name': self._clean(row[COL['name'] - 1]),
                    'phone': self._clean(row[COL['phone'] - 1]),
                    'source': self._clean(row[COL['source'] - 1]) or 'other',
                    'shipping_cost': row[COL['shipping_cost'] - 1],
                    'lines': []
                }

            # Додаємо позицію
            product_sku = self._clean(row[COL['product_number'] - 1])
            qty = row[COL['qty'] - 1]
            unit_price = row[COL['unit_price'] - 1]
            total_value = row[COL['total_value'] - 1]

            if product_sku and qty:
                orders_data[order_number]['lines'].append({
                    'sku': product_sku,
                    'qty': qty,
                    'unit_price': self._parse_price(unit_price),
                    'total_value': self._parse_price(total_value),
                })

        # Зберігаємо в БД
        for order_number, data in orders_data.items():
            # Перевірка чи вже існує
            if SalesOrder.objects.filter(
                source=data['source'],
                order_number=order_number
            ).exists():
                skipped += 1
                self.stdout.write(f'  ⏭ Вже є: {order_number}')
                continue

            if dry_run:
                self.stdout.write(f'  [DRY] {order_number} ({len(data["lines"])} позицій)')
                created += 1
                continue

            # Створюємо замовлення
            try:
                with transaction.atomic():
                    # Визначаємо статус
                    status = "shipped" if data['shipped_at'] else "received"
                    
                    order = SalesOrder.objects.create(
                        source=data['source'],
                        status=status,
                        order_number=order_number,
                        order_date=self._parse_date(data['order_date'], date_only=True),
                        shipped_at=self._parse_date(data['shipped_at'], date_only=True),
                        shipping_courier=data['courier'],
                        tracking_number=data['tracking'],
                        lieferschein_nr=data['lieferschein'],
                        shipping_region=data['region'],
                        shipping_address=data['address'],
                        shipping_deadline=self._parse_date(data['deadline'], date_only=True),
                        # ВАЖЛИВО: Firma — це назва компанії!
                        client=data['firma'] or data['name'] or data['client'],
                        email=data['email'] if data['email'] not in ['NA', 'N/A', '', None] else '',
                        phone=data['phone'] or '',
                        document_type='SALE',
                        affects_stock=True,
                        shipping_cost=self._parse_price(data.get('shipping_cost')),
                    )

                    # Додаємо позиції
                    for line_data in data['lines']:
                        product = self._find_product(line_data['sku'])
                        
                        SalesOrderLine.objects.create(
                            order=order,
                            product=product,
                            sku_raw=line_data['sku'],
                            qty=Decimal(str(line_data['qty'])),
                            unit_price=line_data['unit_price'],
                            total_price=line_data['total_value'],
                        )

                    created += 1
                    self.stdout.write(self.style.SUCCESS(
                        f'  ✅ {order_number} ({len(data["lines"])} позицій)'
                    ))

            except Exception as e:
                self.stderr.write(self.style.ERROR(
                    f'  ❌ Помилка {order_number}: {e}'
                ))

        return created, skipped

    def _clean(self, value):
        """Очищає значення."""
        if value is None:
            return ''
        s = str(value).strip()
        if s.upper() in ['NA', 'N/A', 'NONE', '-']:
            return ''
        return s

    def _parse_date(self, value, date_only=False):
        """Парсить дату."""
        if not value:
            return None
        if isinstance(value, datetime):
            if date_only:
                return value.date()
            return timezone.make_aware(value) if timezone.is_naive(value) else value
        # Спроба парсити рядок
        try:
            from dateutil import parser
            dt = parser.parse(str(value))
            if date_only:
                return dt.date()
            return timezone.make_aware(dt) if timezone.is_naive(dt) else dt
        except Exception:
            return None

    def _parse_price(self, value):
        """Парсить ціну. Підтримує European (37,82) і US (1,234.56) формати."""
        if not value:
            return Decimal('0')
        s = str(value).strip()
        # Strip non-numeric except comma and period
        import re as _re
        price_str = _re.sub(r'[^\d\.,]', '', s)
        if not price_str:
            return Decimal('0')
        # Determine decimal separator: if both present, last one wins
        has_comma = ',' in price_str
        has_dot   = '.' in price_str
        if has_comma and has_dot:
            if price_str.rfind(',') > price_str.rfind('.'):
                # European: 1.234,56
                price_str = price_str.replace('.', '').replace(',', '.')
            else:
                # US: 1,234.56
                price_str = price_str.replace(',', '')
        else:
            price_str = price_str.replace(',', '.')
        try:
            return Decimal(price_str) if price_str else Decimal('0')
        except Exception:
            return Decimal('0')

    def _find_product(self, sku):
        """Знаходить товар по SKU або alias."""
        from inventory.models import Product, ProductAlias
        if not sku:
            return None
        product = Product.objects.filter(sku__iexact=sku).first()
        if product:
            return product
        alias = ProductAlias.objects.filter(alias__iexact=sku).select_related('product').first()
        return alias.product if alias else None
