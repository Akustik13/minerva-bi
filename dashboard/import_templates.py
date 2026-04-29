"""
Excel import template generators for Inventory and Sales wizards.
Called from dashboard/urls.py download view.

Structure of each template:
  Sheet 1 ("Products" / "Orders"):
    Row 1  — column headers  ← wizard auto-detects this as header row
    Rows 2+ — example data   ← wizard imports from here
  Sheet 2 ("Інструкція"):
    Full field documentation, allowed values, tips
"""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── helpers ────────────────────────────────────────────────────────────────────

def _side(style="thin", color="CBD5E0"):
    return Side(style=style, color=color)


_BORDER    = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
_THICK_BOT = Border(left=_side(), right=_side(), top=_side(),
                    bottom=Side(style="medium", color="94A3B8"))


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _font(bold=False, italic=False, size=10, color="1E293B") -> Font:
    return Font(name="Calibri", bold=bold, italic=italic, size=size, color=color)


def _align(wrap=False, h="left", v="center") -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _col_letter(n): return get_column_letter(n)


# ═══════════════════════════════════════════════════════════════════════════════
# INVENTORY TEMPLATE
# ═══════════════════════════════════════════════════════════════════════════════

# (field_name, label_for_docs, col_width, required, ex1, ex2, ex3)
INVENTORY_COLS = [
    ("sku",            "Артикул",          16, True,
     "RF-001",        "CABLE-2M",          "ANT-UHF-5dB"),
    ("name",           "Назва",             28, False,
     "RF фільтр 433 МГц", "Кабель коаксіальний 2 м", "Антена UHF 5 dBi"),
    ("category",       "Категорія (slug)",  20, False,
     "rf_filter",     "cable",             "antenna"),
    ("kind",           "Тип",              14, False,
     "finished",      "finished",          "component"),
    ("unit_type",      "Одиниця",          14, False,
     "piece",         "meter",             "piece"),
    ("manufacturer",   "Виробник",         18, False,
     "Murata",        "Belden",            "Taoglas"),
    ("purchase_price", "Ціна закупівлі",   18, False,
     "12.50",         "3.80",              "24.00"),
    ("sale_price",     "Ціна продажу",     16, False,
     "19.90",         "6.50",             "38.00"),
    ("reorder_point",  "Поріг reorder",    16, False,
     "5",             "20",               "3"),
    ("lead_time_days", "Термін постачання",18, False,
     "14",            "7",                "21"),
    ("initial_stock",  "Початковий залишок",20, False,
     "50",            "200",              "10"),
    ("is_active",      "Активний",         12, False,
     "1",             "1",                "1"),
    ("datasheet_url",  "Datasheet URL",    40, False,
     "https://example.com/ds.pdf", "", "https://taoglas.com/ant.pdf"),
    ("image_url",      "Image URL",        40, False,
     "https://example.com/img.jpg", "", "https://taoglas.com/ant.jpg"),
    ("notes",          "Примітки",         28, False,
     "",              "Стандартний RG-58", "Для зовнішнього монтажу"),
]


def build_inventory_template() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.freeze_panes = "A2"          # freeze header row

    n = len(INVENTORY_COLS)

    # ── Row 1: column headers (wizard reads this as the header row) ────────────
    for col_idx, (fname, _lbl, _w, required, *_) in enumerate(INVENTORY_COLS, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.value = fname                          # exact DB field name
        c.font  = _font(bold=True, size=10, color="FFFFFF")
        c.fill  = _fill("1B5E20") if required else _fill("2E7D32")
        c.alignment = _align(h="center")
        c.border = _THICK_BOT
    ws.row_dimensions[1].height = 22

    # ── Row 1 comment strip (merged, below headers — visible hint) ─────────────
    # We don't add a hint row here because the wizard imports ALL rows after row 1.
    # Hints are on the "Інструкція" sheet instead.

    # ── Rows 2-4: example data ─────────────────────────────────────────────────
    ROW_FILLS = [_fill("F1F8E9"), _fill("FFFFFF"), _fill("F1F8E9")]
    for row_offset, rfill in enumerate(ROW_FILLS):
        row = 2 + row_offset
        for col_idx, col in enumerate(INVENTORY_COLS, start=1):
            ex_val = col[4 + row_offset]           # ex1, ex2, ex3
            c = ws.cell(row=row, column=col_idx)
            c.value = ex_val
            c.font  = _font(size=10)
            c.fill  = rfill
            c.alignment = _align()
            c.border = _BORDER
        ws.row_dimensions[row].height = 18

    # ── Column widths ──────────────────────────────────────────────────────────
    for col_idx, col in enumerate(INVENTORY_COLS, start=1):
        ws.column_dimensions[_col_letter(col_idx)].width = col[2]

    # ── Sheet 2: instructions ──────────────────────────────────────────────────
    inst = wb.create_sheet("\u0406\u043d\u0441\u0442\u0440\u0443\u043a\u0446\u0456\u044f")  # "Інструкція"
    _write_inventory_instructions(inst)
    return wb


def _write_inventory_instructions(ws):
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30

    rows = [
        # (col A, col B, col C, col D)
        ("Поле",           "Опис",                                  "Обов'язкове", "Допустимі значення / Формат"),
        ("sku",            "Унікальний артикул. Регістр важливий.",  "★ Так",       "Будь-який рядок. Напр.: RF-001"),
        ("name",           "Назва товару / опис.",                  "Ні",          "Будь-який текст"),
        ("category",       "Slug категорії з довідника.",           "Ні",          "antenna / cable / rf_filter / other"),
        ("kind",           "Тип виробу.",                           "Ні",          "finished  або  component"),
        ("unit_type",      "Одиниця вимірювання.",                  "Ні",          "piece / meter / kilogram / liter / set"),
        ("manufacturer",   "Виробник.",                             "Ні",          "Будь-який текст. Напр.: Murata"),
        ("purchase_price", "Ціна закупівлі за одиницю.",            "Ні",          "12.50  або  12,50  (кома замінюється)"),
        ("sale_price",     "Ціна продажу за одиницю.",              "Ні",          "19.90  або  19,90"),
        ("reorder_point",  "Мінімальний залишок для повторного\nзамовлення. Ціле число ≥ 0.", "Ні", "5"),
        ("lead_time_days", "Стандартний термін постачання в днях.", "Ні",          "14"),
        ("initial_stock",  "Початковий залишок.\nСистема автоматично створить\nтранзакцію типу ADJUSTMENT.", "Ні", "50"),
        ("is_active",      "Чи активний товар.",                    "Ні",          "1 / true / yes → активний\n0 / false / no → неактивний"),
        ("notes",          "Довільні нотатки.",                     "Ні",          "Будь-який текст"),
        ("", "", "", ""),
        ("ВАЖЛИВО",        "Рядок 1 — назви полів (заголовок).\nРядки 2+ — дані для імпорту.\nНЕ додавайте зайвих рядків перед даними.", "", ""),
        ("Режим Skip",     "Якщо SKU вже є в БД — рядок пропускається.",  "", ""),
        ("Режим Update",   "Якщо SKU вже є в БД — поля оновлюються.",     "", ""),
        ("Dry-run",        "Тест без змін у БД. Рекомендовано для перевірки.", "", ""),
    ]

    for r_idx, (a, b, c, d) in enumerate(rows, start=1):
        is_hdr = r_idx == 1
        is_note = r_idx >= 16
        for c_idx, val in enumerate([a, b, c, d], start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val
            cell.font = _font(bold=is_hdr or is_note, size=10,
                              color="FFFFFF" if is_hdr else ("C8E6C9" if is_note else "1E293B"))
            cell.fill = _fill("1B5E20") if is_hdr else (
                _fill("1A2535") if is_note else (
                _fill("F1F8E9") if r_idx % 2 == 0 else _fill("FFFFFF")))
            cell.alignment = _align(wrap=True, v="top")
            cell.border = _BORDER
        ws.row_dimensions[r_idx].height = 36 if "\n" in b else 18


# ═══════════════════════════════════════════════════════════════════════════════
# SALES TEMPLATE
# ═══════════════════════════════════════════════════════════════════════════════

# (field_name, col_width, required, ex_row1, ex_row2, ex_row3, ex_row4)
SALES_COLS = [
    # ── order-level ──────────────────────────────────────────────────────────
    ("order_number",    18, True,
     "ORD-2024-001", "ORD-2024-001", "ORD-2024-002", "ORD-2024-002"),
    ("source",          14, False,
     "manual",       "manual",       "digikey",       "digikey"),
    ("order_date",      16, False,
     "2024-03-15",   "",             "2024-03-16",    ""),
    ("shipped_at",      16, False,
     "2024-03-18",   "",             "2024-03-20",    ""),
    ("shipping_deadline",18, False,
     "2024-03-17",   "",             "2024-03-19",    ""),
    ("shipping_courier",18, False,
     "DPD",          "",             "GLS",           ""),
    ("tracking_number", 22, False,
     "1Z999AA10123456784", "",       "5318496528237892", ""),
    ("client",          22, False,
     "Acme GmbH",    "",             "SCIENCE CORPORATION", ""),
    ("contact_name",    18, False,
     "Hans Müller",  "",             "ANGELA LEONES", ""),
    ("phone",           16, False,
     "+49 30 123456","",             "",              ""),
    ("email",           22, False,
     "hans@acme.de", "",             "purchasing@science.xyz", ""),
    ("ship_name",       22, False,
     "",             "",             "CHARLES GORDON",""),
    ("ship_company",    22, False,
     "Acme GmbH",    "",             "SCIENCE CORPORATION",""),
    ("ship_phone",      16, False,
     "+49 30 123456","",             "16505615990",   ""),
    ("ship_email",      22, False,
     "",             "",             "",              ""),
    ("addr_street",     22, False,
     "Hauptstraße 42","",            "вул. Хрещатик 22",""),
    ("addr_city",       14, False,
     "Berlin",       "",             "Kyiv",          ""),
    ("addr_zip",        12, False,
     "10115",        "",             "01001",         ""),
    ("addr_country",    12, False,
     "DE",           "",             "UA",            ""),
    ("currency",        10, False,
     "EUR",          "",             "EUR",           ""),
    ("shipping_cost",   14, False,
     "8.50",         "",             "6.00",          ""),
    ("status",          14, False,
     "shipped",      "",             "received",      ""),
    # ── line-level ────────────────────────────────────────────────────────────
    ("sku_raw",         18, True,
     "RF-001",       "CABLE-2M",     "ANT-UHF-5dB",   "RF-001"),
    ("qty",            10, True,
     "2",            "5",            "1",             "3"),
    ("unit_price",      14, False,
     "19.90",        "6.50",         "38.00",         "19.90"),
    ("total_price_line",18, False,
     "39.80",        "32.50",        "38.00",         "59.70"),
]

_SALES_ORDER_FIELDS = {
    "order_number","source","order_date","shipped_at","shipping_deadline",
    "shipping_courier","tracking_number","client","contact_name","phone","email",
    "ship_name","ship_company","ship_phone","ship_email",
    "addr_street","addr_city","addr_zip","addr_country",
    "currency","shipping_cost","status",
}
_SALES_LINE_FIELDS = {"sku_raw","qty","unit_price","total_price_line"}


def build_sales_template() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.freeze_panes = "A2"

    # ── Row 1: column headers ──────────────────────────────────────────────────
    for col_idx, col in enumerate(SALES_COLS, start=1):
        fname, _w, required = col[0], col[1], col[2]
        is_line = fname in _SALES_LINE_FIELDS
        c = ws.cell(row=1, column=col_idx)
        c.value = fname
        c.font  = _font(bold=True, size=10, color="FFFFFF")
        c.fill  = _fill("0D47A1") if required else (
                  _fill("1565C0") if not is_line else _fill("1976D2"))
        c.alignment = _align(h="center")
        c.border = _THICK_BOT
    ws.row_dimensions[1].height = 22

    # ── Rows 2-5: example data (2 orders × 2 lines) ───────────────────────────
    # Order 1 rows: light blue; Order 2 rows: light green
    ORD_FILLS = [
        (_fill("EFF6FF"), _fill("DBEAFE")),   # order 1 line 1 & 2
        (_fill("F0FDF4"), _fill("DCFCE7")),   # order 2 line 1 & 2
    ]
    examples = [col[3:] for col in SALES_COLS]   # 4 example values per col

    for ex_idx in range(4):
        row = 2 + ex_idx
        order_pair = 0 if ex_idx < 2 else 1
        fill_order, fill_line = ORD_FILLS[order_pair]

        for col_idx, col in enumerate(SALES_COLS, start=1):
            fname = col[0]
            c = ws.cell(row=row, column=col_idx)
            c.value = examples[col_idx - 1][ex_idx]
            c.font  = _font(size=10)
            c.fill  = fill_line if fname in _SALES_LINE_FIELDS else fill_order
            c.alignment = _align()
            c.border = _BORDER
        ws.row_dimensions[row].height = 18

    # ── Column widths ──────────────────────────────────────────────────────────
    for col_idx, col in enumerate(SALES_COLS, start=1):
        ws.column_dimensions[_col_letter(col_idx)].width = col[1]

    # ── Sheet 2: instructions ──────────────────────────────────────────────────
    inst = wb.create_sheet("\u0406\u043d\u0441\u0442\u0440\u0443\u043a\u0446\u0456\u044f")
    _write_sales_instructions(inst)
    return wb


def _write_sales_instructions(ws):
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30

    rows = [
        ("Поле",             "Опис",                                       "Обов'язкове", "Формат / Приклад"),
        # order fields
        ("order_number",     "Унікальний номер замовлення.\nРядки з однаковим order_number\nоб'єднуються в одне замовлення.", "★ Так", "ORD-2024-001"),
        ("source",           "Джерело замовлення.",                        "Ні",          "manual / digikey / nova_post"),
        ("order_date",       "Дата замовлення.",                           "Ні",          "2024-03-15\nабо DD.MM.YYYY"),
        ("shipped_at",       "Дата фактичної відправки.",                  "Ні",          "2024-03-18"),
        ("shipping_deadline","Дедлайн відправки.",                         "Ні",          "2024-03-17"),
        ("shipping_courier", "Назва кур'єрської служби.",                  "Ні",          "DPD / GLS / Nova Poshta"),
        ("tracking_number",  "Трекінг-номер кур'єра.",                     "Ні",          "1Z999AA10123456784"),
        ("client",           "Назва компанії або ПІБ клієнта.",            "Ні",          "Acme GmbH"),
        ("contact_name",     "Контактна особа.",                           "Ні",          "Hans Müller"),
        ("phone",            "Номер телефону (будь-який формат).",         "Ні",          "+49 30 12345678"),
        ("email",            "Email клієнта.",                             "Ні",          "hans@acme.de"),
        ("addr_street",      "Вулиця та будинок.",                         "Ні",          "Hauptstraße 42"),
        ("addr_city",        "Місто.",                                     "Ні",          "Berlin / Kyiv"),
        ("addr_zip",         "Поштовий індекс.",                           "Ні",          "10115 / 01001"),
        ("addr_country",     "Код країни ISO-2 (рівно 2 літери).",         "Ні",          "DE / UA / PL / US / CZ"),
        ("currency",         "Валюта замовлення ISO-4217.",                "Ні",          "EUR / USD / GBP"),
        ("shipping_cost",    "Вартість доставки.\nДопускається символ валюти.", "Ні",     "8.50 або $8.50"),
        ("status",           "Статус замовлення.",                         "Ні",          "received / processing\nshipped / cancelled"),
        # line fields
        ("sku_raw",          "SKU або alias товару.\nЯкщо не знайдено в БД —\nствориться placeholder-товар.", "★ Так", "RF-001 / CABLE-2M"),
        ("qty",              "Кількість товару в рядку.\nМає бути > 0.",   "★ Так",       "2 / 5.5"),
        ("unit_price",       "Ціна за одиницю.\nДопускається символ валюти.", "Ні",       "19.90 / €19.90"),
        ("total_price_line", "Сума рядка.\nЯкщо порожньо і є unit_price —\nsystem рахує qty × unit_price.", "Ні", "39.80"),
        ("", "", "", ""),
        ("ВАЖЛИВО",          "Рядок 1 — назви полів (заголовок).\nРядки 2+ — дані для імпорту.\nНЕ додавайте зайвих рядків перед даними.", "", ""),
        ("Багаторядкові",    "Кілька рядків з однаковим order_number\n→ одне замовлення, кілька SKU.\nПоля замовлення можна заповнити\nтільки в першому рядку.", "", ""),
    ]

    for r_idx, (a, b, c, d) in enumerate(rows, start=1):
        is_hdr  = r_idx == 1
        is_note = r_idx >= 25
        is_line = a in _SALES_LINE_FIELDS
        for c_idx, val in enumerate([a, b, c, d], start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val
            cell.font = _font(bold=is_hdr or is_note, size=10,
                              color="FFFFFF" if is_hdr else ("BBDEFB" if is_note else "1E293B"))
            cell.fill = _fill("0D47A1") if is_hdr else (
                _fill("0D2137") if is_note else (
                _fill("E8F4FD") if is_line else (
                _fill("E3F2FD") if r_idx % 2 == 0 else _fill("FFFFFF"))))
            cell.alignment = _align(wrap=True, v="top")
            cell.border = _BORDER
        ws.row_dimensions[r_idx].height = 18 if "\n" not in b else 50
