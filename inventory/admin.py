from __future__ import annotations

from core.mixins import AuditableMixin
from decimal import Decimal, InvalidOperation
import uuid
import json
import os
import tempfile

from django import forms
from django.contrib import admin, messages
from django.db import transaction
from django.db.models import Sum
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from datetime import timedelta

from .forms import ExcelUploadForm, SetStockForm
from .models import (
    Product, ProductAlias, Location,
    InventoryTransaction, ProductComponent,
    Supplier, PurchaseOrder, PurchaseOrderLine,
    ProductCategory, InventorySettings,
)
from shipping.models import ProductPackaging


# ── Утиліти ────────────────────────────────────────────────────────────────────

def _get_stock(product):
    result = InventoryTransaction.objects.filter(
        product=product).aggregate(total=Sum('qty'))
    return float(result['total'] or 0)


def _get_monthly_sales(product, months=3):
    try:
        from sales.models import SalesOrderLine
        since = timezone.now() - timedelta(days=30 * months)
        result = SalesOrderLine.objects.filter(
            product=product,
            order__order_date__gte=since,
        ).aggregate(total=Sum('qty'))
        total = float(result['total'] or 0)
        return total / months
    except Exception:
        return 0


def _reorder(product, stock=None, sales_3m_total=None):
    """Повертає dict з аналізом reorder.
    stock: optional pre-computed stock (from queryset annotation)
    sales_3m_total: optional pre-computed 3-month sales qty (from annotation)
    """
    if stock is None:
        stock = _get_stock(product)
    else:
        stock = float(stock)
    # Від'ємний stock = борг — для розрахунків вважаємо 0
    stock_calc = max(0.0, stock)

    if sales_3m_total is not None:
        monthly = float(sales_3m_total) / 3
    else:
        monthly = _get_monthly_sales(product, months=3)
    if monthly <= 0:
        monthly = _get_monthly_sales(product, months=24)
        if monthly <= 0:
            return {'stock': stock, 'monthly': 0,
                    'months_left': None, 'reorder_qty': 0, 'status': 'no_sales'}
        monthly = monthly / 6

    months_left = stock_calc / monthly if monthly > 0 else 999
    target = monthly * 3 * 1.2
    reorder_qty = max(0, target - stock_calc)
    reorder_qty = round(reorder_qty / 10) * 10 if reorder_qty > 10 else round(reorder_qty)

    if stock <= 0 and monthly > 0: status = 'critical'
    elif months_left <= 0.5:       status = 'critical'
    elif months_left <= 1.5:       status = 'warning'
    elif months_left <= 3:         status = 'low'
    else:                          status = 'ok'

    return {
        'stock': stock, 'monthly': round(monthly, 1),
        'months_left': round(months_left, 1) if stock >= 0 else 0.0,
        'reorder_qty': int(reorder_qty), 'status': status,
    }


# ── Фільтр по статусу ─────────────────────────────────────────────────────────

class ReorderStatusFilter(admin.SimpleListFilter):
    title = "Статус запасів"
    parameter_name = "reorder_status"

    def lookups(self, request, model_admin):
        return [
            ("critical", "🔥 Критично (< 2 тижні)"),
            ("warning",  "⚠️ Мало (< 1.5 міс)"),
            ("low",      "📉 Низько (< 3 міс)"),
            ("ok",       "✅ OK"),
            ("no_sales", "💤 Нема продажів"),
        ]

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        pks = [p.pk for p in queryset if _reorder(p)['status'] == self.value()]
        return queryset.filter(pk__in=pks)


# ── Proxy модель — окрема сторінка аналізу ────────────────────────────────────

class ReorderProxy(Product):
    class Meta:
        proxy = True
        verbose_name        = "Аналіз запасів"
        verbose_name_plural = "📊 Аналіз запасів"


@admin.register(ReorderProxy)
class ReorderAnalysisAdmin(admin.ModelAdmin):

    list_display = (
        "sku", "name_col", "category",
        "stock_col", "monthly_col", "months_left_col",
        "reorder_col", "status_col", "po_btn",
    )
    list_filter  = ("category", "kind", ReorderStatusFilter)
    search_fields = ("sku", "name")
    ordering     = ("sku",)
    actions      = ["bulk_create_po"]

    def has_add_permission(self, request):    return False
    def has_change_permission(self, request, obj=None): return False
    def has_delete_permission(self, request, obj=None): return False

    def get_queryset(self, request):
        from django.db.models import OuterRef, Subquery, ExpressionWrapper, DecimalField, Value
        from django.db.models.functions import Coalesce
        from django.utils import timezone
        from datetime import timedelta
        try:
            from sales.models import SalesOrderLine
            since_3m = (timezone.now() - timedelta(days=90)).date()
            stock_subq = (
                InventoryTransaction.objects.filter(product=OuterRef('pk'))
                .values('product').annotate(t=Sum('qty')).values('t')
            )
            sales_subq = (
                SalesOrderLine.objects.filter(
                    product=OuterRef('pk'),
                    order__order_date__gte=since_3m,
                ).values('product').annotate(t=Sum('qty')).values('t')
            )
            return (
                super().get_queryset(request).filter(is_active=True)
                .annotate(
                    _stock_total=Coalesce(Subquery(stock_subq), Value(Decimal('0'))),
                    _sales_3m=Coalesce(Subquery(sales_subq), Value(Decimal('0'))),
                )
            )
        except Exception:
            return super().get_queryset(request).filter(is_active=True)

    def name_col(self, obj):
        n = obj.name
        return (n[:45] + "…") if len(n) > 45 else n
    name_col.short_description = "Назва"

    def _get_reorder_cached(self, obj):
        if not hasattr(obj, '_reorder_cache'):
            s = getattr(obj, '_stock_total', None)
            s3m = getattr(obj, '_sales_3m', None)
            obj._reorder_cache = _reorder(
                obj,
                stock=float(s) if s is not None else None,
                sales_3m_total=float(s3m) if s3m is not None else None,
            )
        return obj._reorder_cache

    def stock_col(self, obj):
        s = float(getattr(obj, '_stock_total', None) or _get_stock(obj))
        if s <= 0:
            return format_html('<b style="color:#f44336">🚫 0</b>')
        elif s < 5:
            return format_html('<b style="color:#ff9800">⚠️ {}</b>', int(s))
        return format_html('<b style="color:#4caf50">{}</b>', int(s))
    stock_col.short_description = "На складі"

    def monthly_col(self, obj):
        d = self._get_reorder_cached(obj)
        if not d['monthly']:
            return format_html('<span style="opacity:.4">—</span>')
        return format_html('{}/міс', d['monthly'])
    monthly_col.short_description = "Прод./міс"

    def months_left_col(self, obj):
        d = self._get_reorder_cached(obj)
        ml = d['months_left']
        if ml is None:
            return format_html('<span style="opacity:.4">—</span>')
        ml_str = f"{ml:.1f}"
        if ml <= 0.5:
            return format_html('<b style="color:#f44336">🔥 {} міс</b>', ml_str)
        elif ml <= 1.5:
            return format_html('<b style="color:#ff9800">⚡ {} міс</b>', ml_str)
        elif ml <= 3:
            return format_html('<span style="color:#ffb300">{} міс</span>', ml_str)
        return format_html('<span style="color:#4caf50">{} міс</span>', ml_str)
    months_left_col.short_description = "Вистачить"

    def reorder_col(self, obj):
        d = self._get_reorder_cached(obj)
        if not d['reorder_qty']:
            return format_html('<span style="opacity:.4">—</span>')
        return format_html('<b style="color:#2196f3">📦 {} шт.</b>', d['reorder_qty'])
    reorder_col.short_description = "Замовити"

    def status_col(self, obj):
        d = self._get_reorder_cached(obj)
        cfg = {
            'critical': ('#f44336', '🔥 КРИТИЧНО'),
            'warning':  ('#ff9800', '⚠️ Мало'),
            'low':      ('#ffb300', '📉 Низько'),
            'ok':       ('#4caf50', '✅ OK'),
            'no_sales': ('#607d8b', '💤 Нема продажів'),
        }
        color, label = cfg.get(d['status'], ('#607d8b', d['status']))
        return format_html(
            '<span style="background:{};color:#fff;padding:3px 10px;border-radius:12px;'
            'font-size:11px;font-weight:bold;white-space:nowrap">{}</span>',
            color, label)
    status_col.short_description = "Статус"

    def po_btn(self, obj):
        d = _reorder(obj)
        if not d['reorder_qty']:
            return "—"
        url = f"/admin/inventory/purchaseorder/add/?product={obj.pk}&qty={d['reorder_qty']}"
        return format_html(
            '<a href="{}" style="background:#1976d2;color:#fff;padding:4px 12px;'
            'border-radius:6px;text-decoration:none;font-size:11px;white-space:nowrap">'
            '+ PO</a>', url)
    po_btn.short_description = "Дія"

    def bulk_create_po(self, request, queryset):
        """Масово додати всі відібрані товари до одного draft PO."""
        po = PurchaseOrder.objects.filter(status='draft').first()
        if not po:
            supplier = Supplier.objects.first()
            if not supplier:
                self.message_user(request,
                    "❌ Спочатку створіть хоча б одного постачальника.", messages.ERROR)
                return
            po = PurchaseOrder.objects.create(
                supplier=supplier,
                status='draft',
                order_date=timezone.now().date(),
                notes="Auto-created by Reorder Analysis",
            )

        added, skipped = 0, 0
        for product in queryset:
            d = _reorder(product)
            if not d['reorder_qty']:
                skipped += 1
                continue
            line, created = PurchaseOrderLine.objects.get_or_create(
                purchase_order=po,
                product=product,
                defaults={
                    'qty_ordered': d['reorder_qty'],
                    'description': f"Reorder: {product.sku}",
                }
            )
            if not created:
                line.qty_ordered = d['reorder_qty']
                line.save()
            added += 1

        po_url = reverse('admin:inventory_purchaseorder_change', args=[po.pk])
        self.message_user(
            request,
            f"✅ Додано {added} позицій до PO {po} (пропущено: {skipped}). "
            f"Відкрийте: /admin/inventory/purchaseorder/{po.pk}/change/",
            messages.SUCCESS,
        )
    bulk_create_po.short_description = "📦 Додати до PO (чернетка)"


# ── Стандартні admin класи (без змін) ─────────────────────────────────────────

_CATEGORY_TIPS = {
    "slug": (
        "Код (slug) — внутрішній ідентифікатор категорії",
        "Це короткий технічний код, який система зберігає\n"
        "всередині кожного товару замість повної назви.\n\n"
        "Правила:\n"
        "• тільки латинські літери та цифри\n"
        "• замість пробілу — підкреслення _\n"
        "• Приклади: antenna, cable, rf_filter\n\n"
        "Структуру бази даних НЕ змінює — це просто\n"
        "текстове значення, яке зберігається в товарах.\n\n"
        "❗ Якщо після збереження категорії ви вирішите\n"
        "змінити цей код — товари що вже мають старий код\n"
        "перестануть відображатися в цій категорії.\n"
        "Назву і колір змінювати можна вільно."
    ),
    "name": (
        "Назва — відображається в інтерфейсі",
        "Це те, що бачать користувачі в списках та фільтрах.\n"
        "Можна писати будь-якою мовою, з пробілами.\n\n"
        "Змінювати можна вільно — на базу даних не впливає."
    ),
    "color": (
        "Колір бейджу (мітки) у форматі HEX",
        "Визначає колір кольорової мітки категорії.\n\n"
        "Формат: # і 6 символів (цифри 0-9 та літери a-f)\n"
        "Приклади:\n"
        "• #e91e63 — рожевий\n"
        "• #2196f3 — синій\n"
        "• #4caf50 — зелений\n"
        "• #607d8b — сірий\n\n"
        "Змінювати можна вільно — лише зовнішній вигляд."
    ),
    "order": (
        "Порядок відображення в списках та дропдаунах",
        "Менше число — вище в списку.\n\n"
        "Приклади: 1, 2, 3 … або 10, 20, 30\n"
        "(числа з запасом зручніші — легше вставити між ними).\n\n"
        "Змінювати можна вільно — на дані не впливає."
    ),
}


def _mtip(field_key, tips_dict):
    """Повертає mark_safe рядок з tooltip-іконкою (?) для поля."""
    title, body = tips_dict.get(field_key, ("", ""))
    body_html = body.replace("\n", "<br>")
    return mark_safe(
        f'<i class="mtip" aria-label="{title}">'
        f'?<span class="mtip-body"><b>{title}</b><br>{body_html}</span>'
        f'</i>'
    )


@admin.register(ProductCategory)
class ProductCategoryAdmin(admin.ModelAdmin):
    list_display       = ("order", "slug", "name", "color_badge", "customs_hs_code", "customs_description_de", "customs_country_of_origin")
    list_display_links = ("slug",)
    list_editable      = ("order",)
    search_fields      = ("slug", "name")
    ordering           = ("order", "name")

    fieldsets = (
        ("🏷️ Категорія", {
            "fields": (("slug", "name"), ("color", "order")),
        }),
        ("🛃 Митне оформлення (CN23)", {
            "fields": (("customs_hs_code", "customs_country_of_origin"), "customs_description_de"),
            "description": "Використовується в CN23 для всіх товарів категорії. "
                           "Власні значення товару мають пріоритет над значеннями категорії.",
        }),
    )

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        for field_key in ("slug", "name", "color", "order"):
            if field_key in form.base_fields:
                form.base_fields[field_key].help_text = _mtip(field_key, _CATEGORY_TIPS)
        return form

    def color_badge(self, obj):
        return format_html(
            '<span style="background:{};color:#fff;padding:3px 12px;border-radius:10px;'
            'font-size:12px;font-weight:bold">{}</span>',
            obj.color, obj.name)
    color_badge.short_description = "Вигляд бейджу"


@admin.register(ProductAlias)
class ProductAliasAdmin(admin.ModelAdmin):
    list_display  = ("alias", "product")
    search_fields = ("alias", "product__sku")


@admin.register(Location)
class LocationAdmin(admin.ModelAdmin):
    list_display  = ("code", "name")
    search_fields = ("code", "name")


@admin.register(InventoryTransaction)
class InventoryTransactionAdmin(AuditableMixin, admin.ModelAdmin):
    list_display  = ("tx_type", "signed_qty", "product", "location",
                     "ref_doc", "tx_date", "created_at", "performer_col")
    search_fields = ("product__sku", "ref_doc", "external_key")
    list_filter   = ("tx_type", "location__code", "product__category")
    date_hierarchy = "created_at"

    def signed_qty(self, obj):
        try:
            q = Decimal(str(obj.qty))
            return f"+{q}" if q > 0 else str(q)
        except Exception:
            return obj.qty
    signed_qty.short_description = "Qty"

    def save_model(self, request, obj, form, change):
        if not getattr(obj, "external_key", None):
            obj.external_key = f"manual:{uuid.uuid4()}"
        super().save_model(request, obj, form, change)

    def changelist_view(self, request, extra_context=None):
        self._performer_map = {}
        response = super().changelist_view(request, extra_context)
        try:
            cl = response.context_data.get('cl')
            pks = [obj.pk for obj in cl.queryset]
            if pks:
                from core.models import AuditLog
                from django.contrib.contenttypes.models import ContentType
                ct = ContentType.objects.get_for_model(InventoryTransaction)
                for entry in AuditLog.objects.filter(
                    content_type=ct,
                    object_id__in=[str(pk) for pk in pks],
                    action='create',
                ).values('object_id', 'user__username', 'user__first_name', 'user__last_name'):
                    uname = (
                        f"{entry['user__first_name']} {entry['user__last_name']}".strip()
                        or entry['user__username'] or ''
                    )
                    self._performer_map[int(entry['object_id'])] = uname
        except Exception:
            pass
        return response

    @admin.display(description="Виконавець")
    def performer_col(self, obj):
        name = getattr(self, '_performer_map', {}).get(obj.pk, '')
        if name:
            return format_html('<span style="color:#42a5f5;font-weight:600">{}</span>', name)
        return format_html('<span style="color:var(--text-dim,#607d8b);font-size:11px">⚙️ Система</span>')


class ProductComponentInline(admin.TabularInline):
    model = ProductComponent
    fk_name = "parent"
    extra = 0
    autocomplete_fields = ("component",)
    fields = ("component", "qty_per", "optional", "note")
    verbose_name_plural = "Components (BOM)"


class ProductPackagingInline(admin.TabularInline):
    model = ProductPackaging
    extra = 0
    fields = ('packaging', 'qty_per_box', 'estimated_weight_g', 'is_default', 'notes')
    verbose_name        = 'Рекомендована упаковка'
    verbose_name_plural = '📦 Рекомендована упаковка'


@admin.register(Product)
class ProductAdmin(AuditableMixin, admin.ModelAdmin):
    change_list_template = "admin/inventory/product/change_list.html"
    list_display = (
        "sku", "sku_short", "category", "kind", "bom_type", "is_active",
        "stock_qty", "stock_badge", "incoming_qty", "buildable_qty",
        "reorder_badge", "label_btn", "set_stock_link",
    )
    search_fields = ("sku", "sku_short", "name")
    list_filter   = ("category", "kind", "bom_type", "is_active")
    list_per_page = 50
    inlines       = (ProductComponentInline, ProductPackagingInline)
    readonly_fields = ("stock_qty", "incoming_qty", "buildable_qty",
                       "set_stock_link", "reorder_info", "label_detail",
                       "image_preview", "datasheet_link", "movement_history")
    fieldsets = (
        (None, {"fields": ("sku", "sku_short", "name", "category",
                            "kind", "bom_type", "unit_type", "is_active")}),
        ("💰 Ціни та закупівля", {
            "fields": (
                "manufacturer",
                ("purchase_price", "sale_price"),
                ("reorder_point", "lead_time_days"),
            )
        }),
        ("📦 Availability", {"fields": ("stock_qty", "incoming_qty",
                                        "buildable_qty", "reorder_info")}),
        ("🔗 Медіа та документи", {
            "fields": ("datasheet_url", "datasheet_file", "datasheet_link", "image_url", "image", "image_preview"),
            "classes": ("collapse",),
        }),
        ("🛃 Митне оформлення", {
            "fields": ("name_export", ("hs_code", "country_of_origin"), "net_weight_g"),
            "classes": ("collapse",),
            "description": "Використовується для автоматичної митної декларації CN23",
        }),
        ("📦 Рекомендована упаковка", {
            "fields": (),
            "classes": ("collapse",),
            "description": "Прив'яжи упаковку нижче (inline) — буде відображатись як рекомендація на замовленні",
        }),
        ("📋 Рух товару", {
            "fields": ("movement_history",),
            "classes": ("collapse",),
        }),
        ("🏷️ Етикетка DYMO", {"fields": ("label_detail",)}),
        ("⚙️ Quick actions", {"fields": ("set_stock_link",)}),
        ("📝 Notes",          {"fields": ("notes",)}),
    )

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        cats = list(ProductCategory.objects.order_by("order", "name").values_list("slug", "name"))
        if cats:
            form.base_fields["category"].widget = forms.Select(
                choices=[("", "— оберіть категорію —")] + cats,
                attrs={"style": "max-width:300px"},
            )
        return form

    def get_queryset(self, request):
        """Annotate stock, incoming and 3-month sales totals to avoid N+1 queries."""
        from django.db.models import OuterRef, Subquery, ExpressionWrapper, DecimalField, Value
        from django.db.models.functions import Coalesce
        from django.utils import timezone
        from datetime import timedelta
        try:
            from sales.models import SalesOrderLine
            since_3m = (timezone.now() - timedelta(days=90)).date()
            stock_subq = (
                InventoryTransaction.objects.filter(product=OuterRef('pk'))
                .values('product').annotate(t=Sum('qty')).values('t')
            )
            incoming_subq = (
                PurchaseOrderLine.objects
                .filter(product=OuterRef('pk'),
                        purchase_order__status__in=['draft', 'ordered', 'partial'])
                .values('product')
                .annotate(i=ExpressionWrapper(
                    Sum('qty_ordered') - Sum('qty_received'),
                    output_field=DecimalField(max_digits=18, decimal_places=3),
                ))
                .values('i')
            )
            sales_subq = (
                SalesOrderLine.objects.filter(
                    product=OuterRef('pk'),
                    order__order_date__gte=since_3m,
                ).values('product').annotate(t=Sum('qty')).values('t')
            )
            return super().get_queryset(request).annotate(
                _stock_total=Coalesce(Subquery(stock_subq), Value(Decimal('0'))),
                _incoming_total=Coalesce(Subquery(incoming_subq), Value(Decimal('0'))),
                _sales_3m=Coalesce(Subquery(sales_subq), Value(Decimal('0'))),
            )
        except Exception:
            return super().get_queryset(request)

    def changelist_view(self, request, extra_context=None):
        """Pre-load labels dir listing once per request to avoid per-row filesystem scans."""
        from pathlib import Path
        from django.conf import settings
        labels_dir = Path(getattr(settings, 'LABELS_DIR', Path(settings.BASE_DIR) / 'labels'))
        try:
            self._cached_labels = {f.stem.upper() for f in labels_dir.glob('*.dymo')}
        except Exception:
            self._cached_labels = set()
        return super().changelist_view(request, extra_context)

    # ── Computed columns ──────────────────────────────────────────────────────

    def stock_qty(self, obj):
        if hasattr(obj, '_stock_total'):
            return obj._stock_total
        total = (InventoryTransaction.objects
                 .filter(product=obj).aggregate(total=Sum("qty")).get("total"))
        return total or Decimal("0")
    stock_qty.short_description = "On stock"

    def _get_reorder_cached(self, obj):
        """Return _reorder() result, cached on obj to avoid duplicate calls per row."""
        if not hasattr(obj, '_reorder_cache'):
            s = getattr(obj, '_stock_total', None)
            s3m = getattr(obj, '_sales_3m', None)
            obj._reorder_cache = _reorder(
                obj,
                stock=float(s) if s is not None else None,
                sales_3m_total=float(s3m) if s3m is not None else None,
            )
        return obj._reorder_cache

    def stock_badge(self, obj):
        """Візуальний індикатор залишку."""
        d = self._get_reorder_cached(obj)
        s = d['stock']
        status = d['status']
        if s <= 0:
            return format_html('<span style="color:#f44336;font-weight:bold">🚫</span>')
        if status == 'critical':
            return format_html('<span style="color:#f44336">🔥</span>')
        if status == 'warning':
            return format_html('<span style="color:#ff9800">⚠️</span>')
        if status == 'low':
            return format_html('<span style="color:#ffb300">📉</span>')
        return format_html('<span style="color:#4caf50">✅</span>')
    stock_badge.short_description = "⚡"

    def reorder_badge(self, obj):
        d = self._get_reorder_cached(obj)
        if d['status'] in ('ok', 'no_sales'):
            return format_html('<span style="opacity:.4">—</span>')
        colors = {'critical': '#f44336', 'warning': '#ff9800', 'low': '#ffb300'}
        color = colors.get(d['status'], '#607d8b')
        return format_html(
            '<span style="color:{};font-weight:bold">📦 {}</span>',
            color, d['reorder_qty'])
    reorder_badge.short_description = "Reorder"

    def reorder_info(self, obj):
        """Детальна аналітика в картці товару."""
        d = _reorder(obj)
        status_cfg = {
            'critical': ('#f44336', '🔥 КРИТИЧНО — замовляйте ЗАРАЗ!'),
            'warning':  ('#ff9800', '⚠️ Мало — треба замовити'),
            'low':      ('#ffb300', '📉 Низько — скоро потрібно замовити'),
            'ok':       ('#4caf50', '✅ Запасів достатньо'),
            'no_sales': ('#607d8b', '💤 Продажів не було (немає даних)'),
        }
        color, msg = status_cfg.get(d['status'], ('#607d8b', '—'))

        rows = [
            ("На складі зараз", f"<b>{int(d['stock'])} шт.</b>"),
            ("Середні продажі", f"{d['monthly']}/міс (останні 3 міс)" if d['monthly'] else "—"),
            ("Вистачить на", f"{d['months_left']} міс" if d['months_left'] is not None else "—"),
            ("Рекомендовано замовити", f"<b style='color:#2196f3'>{d['reorder_qty']} шт.</b>" if d['reorder_qty'] else "Не потрібно"),
        ]
        table_rows = "".join(
            f"<tr><td style='padding:6px 12px;opacity:.7;white-space:nowrap'>{k}</td>"
            f"<td style='padding:6px 12px'>{v}</td></tr>"
            for k, v in rows
        )
        return mark_safe(
            f'<div style="border-left:4px solid {color};padding:8px 14px;'
            f'margin-bottom:10px;font-weight:bold;color:{color}">{msg}</div>'
            f'<table style="border-collapse:collapse;border:1px solid rgba(128,128,128,.2);'
            f'border-radius:6px">{table_rows}</table>'
        )
    reorder_info.short_description = "📊 Аналіз запасів"

    def incoming_qty(self, obj):
        if hasattr(obj, '_incoming_total'):
            return obj._incoming_total
        q = (PurchaseOrderLine.objects
             .filter(product=obj,
                     purchase_order__status__in=['draft', 'ordered', 'partial'])
             .aggregate(inc=Sum("qty_ordered") - Sum("qty_received"))
             .get("inc"))
        return q or Decimal("0")
    incoming_qty.short_description = "Incoming"

    def buildable_qty(self, obj):
        if obj.bom_type == Product.BomType.NONE:
            return "-"
        comps = list(ProductComponent.objects.filter(parent=obj).select_related("component"))
        if not comps:
            return "0"
        build = None
        for c in comps:
            if c.optional:
                continue
            cs = float(self.stock_qty(c.component))
            possible = int(Decimal(str(cs)) // Decimal(str(c.qty_per or 1)))
            build = possible if build is None else min(build, possible)
        return build if build is not None else "0"
    buildable_qty.short_description = "Buildable"

    def label_detail(self, obj):
        from pathlib import Path
        from django.conf import settings
        sku = (obj.sku or "").strip()
        if not sku:
            return mark_safe('<p style="color:#607d8b">SKU не вказано</p>')
        labels_dir = Path(getattr(settings, 'LABELS_DIR', Path(settings.BASE_DIR) / 'labels'))
        sku_up = sku.upper()
        label_path = None
        for f in labels_dir.glob('*.dymo'):
            s = f.stem.upper()
            if s == sku_up:
                label_path = f
                break
            if label_path is None and s.startswith(sku_up) and len(s) > len(sku_up) and s[len(sku_up)] in (' ', '_'):
                label_path = f
        upload_js = f"""(function(f){{
  var fd=new FormData(); fd.append('labels',f);
  fetch('/labels/upload/',{{method:'POST',body:fd,
    headers:{{'X-CSRFToken':document.cookie.match(/csrftoken=([^;]+)/)[1]}}}})
  .then(r=>r.json()).then(d=>{{
    if(d.results&&d.results[0]&&d.results[0].status!='error') location.reload();
    else alert('Помилка: '+(d.results?d.results[0].msg:d.error));
  }});
}})(this.files[0])"""
        upload_btn = f'''<label style="display:inline-block;padding:6px 14px;background:#1a2e4a;
    border:1px solid #2a4a6a;border-radius:7px;font-size:12px;color:#58a6ff;
    cursor:pointer;white-space:nowrap">
  <input type="file" accept=".dymo" style="display:none" onchange="{upload_js}">
  📤 {'Замінити файл' if label_path else f'Завантажити {sku}.dymo'}
</label>'''

        if label_path:
            url_base = f"/labels/serve/{sku}/"
            return mark_safe(f'''
<div style="display:flex;flex-direction:column;gap:12px;padding:4px 0">
  <div style="display:flex;align-items:center;gap:16px;flex-wrap:wrap">
    <span style="color:#4caf50;font-size:13px">✅ {label_path.name}</span>
    <div style="display:flex;align-items:center;gap:8px">
      <label style="font-size:12px;color:#9aafbe">К-сть:</label>
      <input id="lbl-qty-{sku}" type="number" value="1" min="1" max="999"
             style="width:60px;padding:4px 6px;background:#111c26;border:1px solid #2a3f52;
                    border-radius:5px;color:#c9d8e4;font-size:13px;text-align:center">
      <a id="lbl-btn-{sku}" href="{url_base}?qty=1" target="_blank"
         style="background:#1976d2;color:#fff;padding:7px 18px;border-radius:7px;
                font-size:13px;font-weight:600;text-decoration:none;white-space:nowrap">
        🖨️ Друкувати
      </a>
    </div>
  </div>
  <div>{upload_btn}</div>
</div>
<script>
(function(){{
  var inp = document.getElementById('lbl-qty-{sku}');
  var btn = document.getElementById('lbl-btn-{sku}');
  inp.addEventListener('input', function(){{
    btn.href = '{url_base}?qty=' + (parseInt(inp.value) || 1);
  }});
}})();
</script>''')

        return mark_safe(f'''
<div style="display:flex;align-items:center;gap:16px;flex-wrap:wrap;padding:4px 0">
  <span style="color:#f44336;font-size:13px">❌ Файл {sku}.dymo не знайдено</span>
  {upload_btn}
</div>''')
    label_detail.short_description = "Друк етикетки"

    def label_btn(self, obj):
        sku = (obj.sku or "").strip()
        if not sku:
            return format_html('<span style="color:#607d8b">—</span>')
        labels = getattr(self, '_cached_labels', None)
        if labels is None:
            from pathlib import Path
            from django.conf import settings
            labels_dir = Path(getattr(settings, 'LABELS_DIR', Path(settings.BASE_DIR) / 'labels'))
            try:
                labels = {f.stem.upper() for f in labels_dir.glob('*.dymo')}
            except Exception:
                labels = set()
        sku_up = sku.upper()
        found = any(
            s == sku_up or (s.startswith(sku_up) and len(s) > len(sku_up) and s[len(sku_up)] in (' ', '_'))
            for s in labels
        )
        if found:
            url = f"/labels/serve/{sku}/?qty=1"
            return mark_safe(
                f'<a href="{url}" '
                f'style="display:inline-block;background:#1976d2;color:#fff;'
                f'padding:3px 10px;border-radius:6px;font-size:11px;text-decoration:none;white-space:nowrap">'
                f'🖨️ Друк</a>')
        return mark_safe(
            f'<span style="color:#607d8b;font-size:11px" title="Немає {sku}.dymo">—</span>')
    label_btn.short_description = "🏷️ Друк"

    def movement_history(self, obj):
        """Рух товару: всі InventoryTransaction з посиланнями на замовлення/PO."""
        if not obj.pk:
            return mark_safe('<span style="color:var(--text-dim,#607d8b)">Збережіть товар спочатку</span>')

        from django.utils.timezone import localtime as _localtime
        txs = list(
            InventoryTransaction.objects.filter(product=obj)
            .select_related('location')
            .order_by('-tx_date', '-pk')[:150]
        )
        if not txs:
            return mark_safe('<span style="color:var(--text-dim,#607d8b)">Транзакцій немає</span>')

        # ── Batch-fetch AuditLog performer ─────────────────────────────────────
        audit_user_map: dict = {}
        try:
            from core.models import AuditLog
            from django.contrib.contenttypes.models import ContentType
            inv_ct = ContentType.objects.get_for_model(InventoryTransaction)
            for entry in AuditLog.objects.filter(
                content_type=inv_ct,
                object_id__in=[str(tx.pk) for tx in txs],
                action='create',
            ).values('object_id', 'user__username', 'user__first_name', 'user__last_name'):
                uname = (
                    f"{entry['user__first_name']} {entry['user__last_name']}".strip()
                    or entry['user__username'] or ''
                )
                audit_user_map[int(entry['object_id'])] = uname
        except Exception:
            pass

        # ── Batch-fetch related SalesOrders ────────────────────────────────────
        so_keys = set()
        po_codes = set()
        for tx in txs:
            ek = tx.external_key
            if ek.startswith('so:'):
                parts = ek.split(':')
                if len(parts) >= 3:
                    so_keys.add((parts[1], parts[2]))
            elif ek.startswith('po:'):
                parts = ek.split(':')
                if len(parts) >= 2:
                    po_codes.add(parts[1])

        orders_map = {}
        if so_keys:
            from sales.models import SalesOrder
            from django.db.models import Q as _Q
            q = _Q()
            for source, order_num in so_keys:
                q |= _Q(source=source, order_number=order_num)
            for o in SalesOrder.objects.filter(q).values(
                'source', 'order_number', 'pk', 'client', 'contact_name'
            ):
                orders_map[(o['source'], o['order_number'])] = o

        po_map = {}
        if po_codes:
            for po in PurchaseOrder.objects.filter(code__in=po_codes).values(
                'code', 'pk', 'supplier__name'
            ):
                po_map[po['code']] = po

        # ── Render ─────────────────────────────────────────────────────────────
        TYPE_CFG = {
            'Incoming':   ('▲', '#4caf50', 'Прихід'),
            'Outgoing':   ('▼', '#ef5350', 'Відвантаження'),
            'Adjustment': ('⇄', '#2196f3', 'Коригування'),
        }

        rows = []
        for tx in txs:
            icon, color, label = TYPE_CFG.get(tx.tx_type, ('?', '#607d8b', tx.tx_type))
            qty = float(tx.qty)
            if tx.tx_type == 'Incoming':
                qty_str = f'+{abs(qty):g}'
            elif tx.tx_type == 'Outgoing':
                qty_str = f'−{abs(qty):g}'
            else:
                qty_str = f'{qty:+g}'

            doc_html = who_html = ''
            ek = tx.external_key

            if ek.startswith('so:'):
                parts = ek.split(':')
                source = parts[1] if len(parts) > 1 else ''
                order_num = parts[2] if len(parts) > 2 else ''
                order = orders_map.get((source, order_num))
                if order:
                    url = f'/admin/sales/salesorder/{order["pk"]}/change/'
                    doc_html = (
                        f'<a href="{url}" style="color:#58a6ff;text-decoration:none">'
                        f'#{order_num}</a>'
                        + (f' <span style="font-size:10px;opacity:.55">[{source}]</span>' if source else '')
                    )
                    client = (order['contact_name'] or order['client'] or '').strip()
                    who_html = f'<span style="font-size:11px">{client[:35]}</span>' if client else ''
                else:
                    doc_html = f'<span style="font-size:11px">{tx.ref_doc}</span>'

            elif ek.startswith('po:'):
                parts = ek.split(':')
                po_code = parts[1] if len(parts) > 1 else ''
                po = po_map.get(po_code)
                if po:
                    url = f'/admin/inventory/purchaseorder/{po["pk"]}/change/'
                    doc_html = f'<a href="{url}" style="color:#58a6ff;text-decoration:none">{po_code}</a>'
                    supplier = (po.get('supplier__name') or '').strip()
                    who_html = f'<span style="font-size:11px">{supplier[:35]}</span>' if supplier else ''
                else:
                    doc_html = f'<span style="font-size:11px">{tx.ref_doc}</span>'

            else:
                doc_html = f'<span style="font-size:11px;opacity:.65">{tx.ref_doc or ek[:40]}</span>'

            date_str = _localtime(tx.tx_date).strftime('%d.%m.%Y %H:%M') if tx.tx_date else '—'
            loc = tx.location.code if tx.location_id else '—'
            _op = audit_user_map.get(tx.pk, '')
            if _op:
                op_html = f'<span style="font-size:11px;color:#42a5f5;font-weight:600">{_op}</span>'
            else:
                op_html = '<span style="font-size:10px;color:var(--text-dim,#607d8b)">⚙️ Система</span>'

            rows.append(
                f'<tr style="border-bottom:1px solid rgba(128,128,128,.1)">'
                f'<td style="padding:5px 10px;white-space:nowrap;font-size:12px;color:var(--text-muted,#9aafbe)">{date_str}</td>'
                f'<td style="padding:5px 10px">'
                f'<span style="background:{color}22;color:{color};padding:2px 8px;border-radius:8px;'
                f'font-size:11px;font-weight:700;white-space:nowrap">{icon} {label}</span></td>'
                f'<td style="padding:5px 10px;text-align:right">'
                f'<b style="color:{color};font-size:13px;font-variant-numeric:tabular-nums;white-space:nowrap">{qty_str}</b></td>'
                f'<td style="padding:5px 10px">{doc_html}</td>'
                f'<td style="padding:5px 10px">{who_html}</td>'
                f'<td style="padding:5px 10px">{op_html}</td>'
                f'<td style="padding:5px 10px;font-size:11px;color:var(--text-muted,#9aafbe)">{loc}</td>'
                f'</tr>'
            )

        total = InventoryTransaction.objects.filter(product=obj).count()
        note = (
            f' <span style="opacity:.5">(показано {len(txs)} з {total})</span>'
            if total > len(txs) else ''
        )

        return mark_safe(
            f'<div style="overflow-x:auto">'
            f'<div style="font-size:12px;color:var(--text-muted,#9aafbe);margin-bottom:8px">'
            f'Всього записів: <b>{total}</b>{note}</div>'
            f'<table style="width:100%;border-collapse:collapse">'
            f'<thead><tr style="border-bottom:2px solid rgba(128,128,128,.2)">'
            f'<th style="padding:5px 10px;text-align:left;font-size:11px;font-weight:600;'
            f'white-space:nowrap;color:var(--text-muted,#9aafbe)">Дата</th>'
            f'<th style="padding:5px 10px;text-align:left;font-size:11px;font-weight:600;'
            f'color:var(--text-muted,#9aafbe)">Тип</th>'
            f'<th style="padding:5px 10px;text-align:right;font-size:11px;font-weight:600;'
            f'color:var(--text-muted,#9aafbe)">К-сть</th>'
            f'<th style="padding:5px 10px;text-align:left;font-size:11px;font-weight:600;'
            f'color:var(--text-muted,#9aafbe)">Документ</th>'
            f'<th style="padding:5px 10px;text-align:left;font-size:11px;font-weight:600;'
            f'color:var(--text-muted,#9aafbe)">Клієнт / Постачальник</th>'
            f'<th style="padding:5px 10px;text-align:left;font-size:11px;font-weight:600;'
            f'color:var(--text-muted,#9aafbe)">Виконавець</th>'
            f'<th style="padding:5px 10px;text-align:left;font-size:11px;font-weight:600;'
            f'color:var(--text-muted,#9aafbe)">Локація</th>'
            f'</tr></thead>'
            f'<tbody>{"".join(rows)}</tbody>'
            f'</table></div>'
        )
    movement_history.short_description = "📋 Рух товару"

    def datasheet_link(self, obj):
        url = obj.datasheet_display_url
        if not url:
            return mark_safe('<span style="color:var(--text-dim,#607d8b);font-size:12px">— не вказано —</span>')
        label = '📎 Відкрити (файл)' if (obj.datasheet_file and obj.datasheet_file.name) else '📄 Відкрити Datasheet'
        return format_html(
            '<a href="{}" target="_blank" rel="noopener noreferrer" '
            'style="display:inline-flex;align-items:center;gap:6px;'
            'background:#1a2e4a;border:1px solid #2a4a6a;border-radius:7px;'
            'padding:6px 14px;font-size:12px;color:#58a6ff;text-decoration:none">'
            '{}</a>',
            url, label,
        )
    datasheet_link.short_description = "Datasheet"

    def image_preview(self, obj):
        url = obj.image.url if obj.image else obj.image_url
        if not url:
            return mark_safe('<span style="color:var(--text-dim,#607d8b);font-size:12px">— немає зображення —</span>')
        return format_html(
            '<a href="{}" target="_blank" rel="noopener noreferrer">'
            '<img src="{}" style="max-height:160px;max-width:320px;border-radius:8px;'
            'border:1px solid rgba(128,128,128,.25);object-fit:contain">'
            '</a>',
            url, url,
        )
    image_preview.short_description = "Попередній перегляд"

    def set_stock_link(self, obj):
        url = reverse("admin:inventory_product_set_stock", args=[obj.pk])
        return format_html('<a class="button" href="{}">Set stock…</a>', url)
    set_stock_link.short_description = "⚙️ Дії"

    def history_view(self, request, object_id, extra_context=None):
        """Рух товару замість стандартного Django LogEntry history."""
        from django.template.response import TemplateResponse
        from django.core.exceptions import PermissionDenied
        from django.db.models import Q as _Q
        from django.utils.timezone import localtime

        obj = get_object_or_404(Product, pk=object_id)
        if not self.has_view_or_change_permission(request, obj):
            raise PermissionDenied

        txs = list(
            InventoryTransaction.objects.filter(product=obj)
            .select_related('location')
            .order_by('-tx_date', '-pk')[:500]
        )

        # ── Batch-fetch SalesOrders and POs ────────────────────────────────────
        so_keys, po_codes = set(), set()
        for tx in txs:
            ek = tx.external_key
            if ek.startswith('so:'):
                parts = ek.split(':')
                if len(parts) >= 3:
                    so_keys.add((parts[1], parts[2]))
            elif ek.startswith('po:'):
                parts = ek.split(':')
                if len(parts) >= 2:
                    po_codes.add(parts[1])

        orders_map = {}
        if so_keys:
            from sales.models import SalesOrder
            q = _Q()
            for source, order_num in so_keys:
                q |= _Q(source=source, order_number=order_num)
            for o in SalesOrder.objects.filter(q).values(
                'source', 'order_number', 'pk', 'client', 'contact_name'
            ):
                orders_map[(o['source'], o['order_number'])] = o

        po_map = {}
        if po_codes:
            for po in PurchaseOrder.objects.filter(code__in=po_codes).values(
                'code', 'pk', 'supplier__name'
            ):
                po_map[po['code']] = po

        # ── Fetch AuditLog creators for manual transactions ─────────────────────
        tx_pk_list = [tx.pk for tx in txs]
        audit_user_map: dict = {}   # tx.pk → username string
        try:
            from core.models import AuditLog
            from django.contrib.contenttypes.models import ContentType
            inv_ct = ContentType.objects.get_for_model(InventoryTransaction)
            for entry in AuditLog.objects.filter(
                content_type=inv_ct,
                object_id__in=[str(pk) for pk in tx_pk_list],
                action='create',
            ).select_related('user').values('object_id', 'user__username', 'user__first_name', 'user__last_name'):
                uid = entry['object_id']
                uname = (
                    f"{entry['user__first_name']} {entry['user__last_name']}".strip()
                    or entry['user__username'] or 'Адмін'
                )
                audit_user_map[int(uid)] = uname
        except Exception:
            pass

        # ── Build rows ─────────────────────────────────────────────────────────
        TYPE_CFG = {
            'Incoming':   ('▲', 'green',  'Прихід'),
            'Outgoing':   ('▼', 'red',    'Відвантаження'),
            'Adjustment': ('⇄', 'blue',   'Коригування'),
        }
        rows = []
        running = 0.0
        for tx in reversed(txs):   # chronological for running total
            running += float(tx.qty)
        running_total = running

        # Re-iterate in display order (newest first) with running balance
        balance = running_total
        for tx in txs:
            kind, color, label = TYPE_CFG.get(tx.tx_type, ('?', 'gray', tx.tx_type))
            qty = float(tx.qty)
            qty_str = (f'+{abs(qty):g}' if tx.tx_type == 'Incoming'
                       else f'−{abs(qty):g}' if tx.tx_type == 'Outgoing'
                       else f'{qty:+g}')
            balance_str = f'{balance:g}'

            doc_url = doc_label = who = ''
            operator = ''   # хто виконав операцію
            ek = tx.external_key
            if ek.startswith('so:'):
                parts = ek.split(':')
                source, order_num = (parts[1] if len(parts) > 1 else ''), (parts[2] if len(parts) > 2 else '')
                order = orders_map.get((source, order_num))
                if order:
                    doc_url   = f'/admin/sales/salesorder/{order["pk"]}/change/'
                    doc_label = f'#{order_num}' + (f' [{source}]' if source else '')
                    who = (order['contact_name'] or order['client'] or '').strip()[:45]
                else:
                    doc_label = tx.ref_doc or ek
                operator = audit_user_map.get(tx.pk, '⚙️ Система')
            elif ek.startswith('po:'):
                parts = ek.split(':')
                po_code = parts[1] if len(parts) > 1 else ''
                po = po_map.get(po_code)
                if po:
                    doc_url   = f'/admin/inventory/purchaseorder/{po["pk"]}/change/'
                    doc_label = po_code
                    who = (po.get('supplier__name') or '').strip()[:45]
                else:
                    doc_label = tx.ref_doc or ek
                operator = audit_user_map.get(tx.pk, '⚙️ Система')
            elif ek.startswith('ai:'):
                doc_label = tx.ref_doc or ek[:55]
                operator = '🤖 AI'
            else:
                doc_label = tx.ref_doc or ek[:55]
                operator = audit_user_map.get(tx.pk, '⚙️ Система')

            rows.append({
                'date':      localtime(tx.tx_date).strftime('%d.%m.%Y %H:%M') if tx.tx_date else '—',
                'label':     label,
                'color':     color,
                'qty_str':   qty_str,
                'balance':   balance_str,
                'doc_url':   doc_url,
                'doc_label': doc_label,
                'who':       who,
                'operator':  operator,
                'loc':       tx.location.code if tx.location_id else '—',
            })
            balance -= qty  # go back in time

        total = InventoryTransaction.objects.filter(product=obj).count()
        ctx = {
            **self.admin_site.each_context(request),
            'title': f'Рух товару — {obj.sku}',
            'object': obj,
            'opts': self.model._meta,
            'rows': rows,
            'total': total,
            'shown': len(txs),
            'current_stock': f'{running_total:g}',
            'has_change_permission': self.has_change_permission(request),
            'preserved_filters': self.get_preserved_filters(request),
        }
        if extra_context:
            ctx.update(extra_context)
        return TemplateResponse(
            request, 'admin/inventory/product/object_history.html', ctx
        )

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path("<path:object_id>/set-stock/",
                 self.admin_site.admin_view(self.set_stock_view),
                 name="inventory_product_set_stock"),
            path("import-excel/",
                 self.admin_site.admin_view(self.import_excel_view),
                 name="inventory_product_import_excel"),
            path("<int:pk>/datasheet/",
                 self.admin_site.admin_view(self.serve_datasheet_view),
                 name="inventory_product_serve_datasheet"),
        ]
        return custom + urls

    def serve_datasheet_view(self, request, pk):
        """Serve product datasheet file via Django (bypasses Synology/nginx media routing)."""
        from django.http import FileResponse, Http404
        product = get_object_or_404(Product, pk=pk)
        if not product.datasheet_file or not product.datasheet_file.name:
            raise Http404('No datasheet file')
        try:
            fname = product.datasheet_file.name.split('/')[-1]
            resp = FileResponse(
                product.datasheet_file.open('rb'),
                content_type='application/pdf',
            )
            resp['Content-Disposition'] = f'inline; filename="{fname}"'
            return resp
        except Exception:
            raise Http404('File not found')

    def set_stock_view(self, request, object_id):
        product = get_object_or_404(Product, pk=object_id)
        current = Decimal(str(self.stock_qty(product)))
        if request.method == "POST":
            form = SetStockForm(request.POST)
            if form.is_valid():
                target = Decimal(str(form.cleaned_data["target_stock"]))
                delta = target - current
                if delta == 0:
                    messages.info(request, "Залишок вже такий — нічого не змінюю.")
                    return redirect(reverse("admin:inventory_product_change", args=[product.pk]))
                location = Location.objects.filter(code="MAIN").first()
                if not location:
                    messages.error(request, "Нема локації MAIN.")
                    return redirect(reverse("admin:inventory_product_change", args=[product.pk]))
                InventoryTransaction.objects.create(
                    external_key=f"manual:setstock:{uuid.uuid4()}",
                    tx_type=InventoryTransaction.TxType.ADJUSTMENT,
                    qty=delta, product=product, location=location,
                    ref_doc="manual:set_stock", tx_date=timezone.now(),
                )
                messages.success(request, f"Було {current}, стало {target}. Adjustment: {delta}.")
                return redirect(reverse("admin:inventory_product_change", args=[product.pk]))
        else:
            form = SetStockForm(initial={"target_stock": current})
        ctx = dict(self.admin_site.each_context(request),
                   title=f"Set stock: {product.sku}", product=product,
                   current_stock=current, form=form,
                   opts=self.model._meta, original=product)
        return render(request, "admin/inventory/product/set_stock.html", ctx)

    # ── Excel Import ───────────────────────────────────────────────────────────

    # Поля БД доступні для маппінгу
    _IMPORT_DB_FIELDS = [
        ("sku",            "SKU (унікальний ключ) *"),
        ("sku_short",      "Короткий SKU"),
        ("name",           "Назва"),
        ("category",       "Категорія (slug з довідника)"),
        ("kind",           "Тип (finished/component)"),
        ("unit_type",      "Одиниця (piece/meter/kilogram/liter/set)"),
        ("manufacturer",   "Виробник"),
        ("purchase_price", "Ціна закупівлі"),
        ("sale_price",     "Ціна продажу"),
        ("reorder_point",  "Поріг reorder"),
        ("lead_time_days", "Термін постачання (днів)"),
        ("datasheet_url",  "Посилання на Datasheet (URL)"),
        ("image_url",      "Посилання на зображення (URL)"),
        ("notes",          "Нотатки"),
        ("is_active",      "Активний (1/0/true/false)"),
        ("initial_stock",  "Початковий залишок (ADJUSTMENT)"),
    ]

    _DECIMAL_FIELDS = {"purchase_price", "sale_price"}
    _INT_FIELDS = {"reorder_point", "lead_time_days"}

    def import_excel_view(self, request):
        """3-крокова форма імпорту товарів з Excel."""
        import openpyxl

        opts = Product._meta
        ctx_base = dict(
            self.admin_site.each_context(request),
            title="📥 Імпорт товарів з Excel",
            opts=opts,
            db_fields=self._IMPORT_DB_FIELDS,
        )

        # ── STEP 1: завантаження файлу ────────────────────────────────────────
        if request.method == "GET" or (request.method == "POST" and request.POST.get("step") not in ("1", "2")):
            form = ExcelUploadForm()
            return render(request, "admin/inventory/import_excel.html",
                          {**ctx_base, "step": 1, "form": form})

        # ── STEP 1 POST: аналіз файлу ─────────────────────────────────────────
        if request.POST.get("step") == "1":
            form = ExcelUploadForm(request.POST, request.FILES)
            if not form.is_valid():
                return render(request, "admin/inventory/import_excel.html",
                              {**ctx_base, "step": 1, "form": form})

            uploaded = request.FILES["excel_file"]
            if not uploaded.name.lower().endswith(".xlsx"):
                form.add_error("excel_file", "Підтримується лише формат .xlsx")
                return render(request, "admin/inventory/import_excel.html",
                              {**ctx_base, "step": 1, "form": form})

            # Зберігаємо тимчасовий файл
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            for chunk in uploaded.chunks():
                tmp.write(chunk)
            tmp.close()
            request.session["excel_import_tmp"] = tmp.name
            request.session["excel_import_name"] = uploaded.name

            # Аналізуємо структуру Excel
            try:
                wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
            except Exception as e:
                os.unlink(tmp.name)
                form.add_error("excel_file", f"Не вдалося відкрити файл: {e}")
                return render(request, "admin/inventory/import_excel.html",
                              {**ctx_base, "step": 1, "form": form})

            # Збираємо назви листів та колонок (перший непустий рядок = заголовок)
            sheet_data = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header_cols = []
                for row in ws.iter_rows(max_row=10, values_only=True):
                    non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if non_empty:
                        header_cols = [str(c).strip() if c is not None else "" for c in row]
                        break
                sheet_data[sheet_name] = header_cols
            wb.close()

            return render(request, "admin/inventory/import_excel.html", {
                **ctx_base,
                "step": 2,
                "sheet_names": wb.sheetnames,
                "sheet_data_json": json.dumps(sheet_data, ensure_ascii=False),
                "first_sheet": wb.sheetnames[0] if wb.sheetnames else "",
                "first_cols": sheet_data.get(wb.sheetnames[0], []) if wb.sheetnames else [],
                "file_name": uploaded.name,
            })

        # ── STEP 2 POST: виконання імпорту ───────────────────────────────────
        if request.POST.get("step") == "2":
            tmp_path = request.session.get("excel_import_tmp")
            if not tmp_path or not os.path.exists(tmp_path):
                messages.error(request, "Сесія застаріла — завантажте файл знову.")
                return redirect(reverse("admin:inventory_product_import_excel"))

            sheet_name   = request.POST.get("sheet_name", "")
            conflict_mode = request.POST.get("conflict_mode", "skip")
            dry_run      = bool(request.POST.get("dry_run"))

            # Збираємо маппінг: col_idx → db_field
            mappings = {}
            for key, val in request.POST.items():
                if key.startswith("col_") and val and val != "--":
                    try:
                        idx = int(key[4:])
                        mappings[idx] = val
                    except ValueError:
                        pass

            if not any(v == "sku" for v in mappings.values()):
                messages.error(request, "Необхідно вибрати колонку для поля SKU.")
                return redirect(reverse("admin:inventory_product_import_excel"))

            # Завантажуємо Excel і запускаємо імпорт
            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]

            stats = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
            location = Location.objects.filter(code="MAIN").first()

            # Пропускаємо рядок заголовка (перший непустий)
            rows_iter = ws.iter_rows(values_only=True)
            header_skipped = False

            with transaction.atomic():
                for row_num, row in enumerate(rows_iter, start=1):
                    # Пропустити порожні рядки
                    if all(c is None or str(c).strip() == "" for c in row):
                        continue
                    # Перший непустий рядок — заголовок, пропускаємо
                    if not header_skipped:
                        header_skipped = True
                        continue

                    # Беремо значення по індексах з маппінгу
                    row_data = {}
                    for idx, field in mappings.items():
                        val = row[idx] if idx < len(row) else None
                        row_data[field] = val

                    sku_raw = row_data.get("sku")
                    if not sku_raw or str(sku_raw).strip() == "":
                        continue
                    sku = str(sku_raw).strip()

                    try:
                        # Готуємо дані для Product
                        product_data = {}
                        for field, val in row_data.items():
                            if field in ("sku", "initial_stock"):
                                continue
                            if val is None:
                                continue
                            sval = str(val).strip()
                            if not sval:
                                continue
                            if field == "is_active":
                                product_data[field] = sval.lower() not in ("0", "false", "ні", "no", "")
                            elif field in self._DECIMAL_FIELDS:
                                try:
                                    product_data[field] = Decimal(sval.replace(",", "."))
                                except InvalidOperation:
                                    pass
                            elif field in self._INT_FIELDS:
                                try:
                                    product_data[field] = int(float(sval.replace(",", ".")))
                                except (ValueError, TypeError):
                                    pass
                            else:
                                product_data[field] = sval

                        exists = Product.objects.filter(sku=sku).exists()

                        if exists and conflict_mode == "skip":
                            stats["skipped"] += 1
                            continue

                        if exists and conflict_mode == "update":
                            Product.objects.filter(sku=sku).update(**product_data)
                            stats["updated"] += 1
                            product = Product.objects.get(sku=sku)
                        else:
                            product, _ = Product.objects.get_or_create(
                                sku=sku, defaults=product_data)
                            stats["created"] += 1

                        # Початковий залишок
                        stock_raw = row_data.get("initial_stock")
                        if stock_raw is not None and str(stock_raw).strip() not in ("", "None"):
                            try:
                                qty = Decimal(str(stock_raw).replace(",", "."))
                                if location and qty != 0:
                                    InventoryTransaction.objects.create(
                                        external_key=f"excel:import:{sku}:{uuid.uuid4()}",
                                        tx_type=InventoryTransaction.TxType.ADJUSTMENT,
                                        qty=qty,
                                        product=product,
                                        location=location,
                                        ref_doc="excel_import",
                                        tx_date=timezone.now(),
                                    )
                            except (InvalidOperation, ValueError):
                                stats["errors"].append(
                                    f"Рядок {row_num}: некоректний залишок '{stock_raw}' для SKU {sku}")

                    except Exception as e:
                        stats["errors"].append(f"Рядок {row_num} (SKU: {sku}): {e}")

                if dry_run:
                    transaction.set_rollback(True)

            wb.close()

            return render(request, "admin/inventory/import_excel.html", {
                **ctx_base,
                "step": 3,
                "stats": stats,
                "dry_run": dry_run,
                "sheet_name": sheet_name,
            })

        # Fallback
        return redirect(reverse("admin:inventory_product_import_excel"))


# ── Inventory Settings ─────────────────────────────────────────────────────────

@admin.register(InventorySettings)
class InventorySettingsAdmin(admin.ModelAdmin):
    fieldsets = (
        ("📤 Списання зі складу (при продажах)", {
            "fields": ("deduct_on", "allow_negative_stock"),
            "description": (
                "<b>При створенні</b> — списується одразу при додаванні замовлення.<br>"
                "<b>При відправці</b> — лише коли статус змінюється на «Відправлено».<br>"
                "<b>При доставці</b> — лише коли статус змінюється на «Доставлено»."
            ),
        }),
        ("📥 Надходження на склад (закупівлі)", {
            "fields": ("add_on_po_receive",),
            "description": "Контролює чи автоматично зараховувати товар при зміні qty_received у PO.",
        }),
        ("📍 Загальні параметри", {
            "fields": ("default_location", "low_stock_alert_enabled"),
        }),
        ("🖼️ Кешування медіа (Datasheet / Фото)", {
            "fields": ("media_cache_panel",),
            "description": (
                "Завантажує PDF-датащити та фото товарів з вказаних URL на сервер. "
                "Після кешування файли відкриваються швидше і залишаються доступними навіть якщо зовнішній сервер недоступний."
            ),
        }),
    )

    readonly_fields = ("media_cache_panel",)

    def media_cache_panel(self, obj):
        total  = Product.objects.count()
        ds_url  = Product.objects.filter(datasheet_url__gt='').count()
        ds_file = Product.objects.exclude(datasheet_file='').exclude(datasheet_file__isnull=True).count()
        ds_need = Product.objects.filter(datasheet_url__gt='').filter(datasheet_file='').count() + \
                  Product.objects.filter(datasheet_url__gt='').filter(datasheet_file__isnull=True).count()
        img_url  = Product.objects.filter(image_url__gt='').count()
        img_file = Product.objects.exclude(image='').exclude(image__isnull=True).count()
        img_need = Product.objects.filter(image_url__gt='').filter(image='').count() + \
                   Product.objects.filter(image_url__gt='').filter(image__isnull=True).count()

        cache_url = reverse('admin:inventory_inventorysettings_media_cache')

        return format_html('''
<div id="inv-mc-panel" style="font-size:13px">
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:14px;max-width:480px">
    <div style="background:var(--bg-card-2,#1e2d40);border:1px solid var(--border-strong);border-radius:6px;padding:10px 14px">
      <div style="font-size:11px;color:var(--text-muted);text-transform:uppercase;letter-spacing:.04em;margin-bottom:6px">📄 Datasheet PDF</div>
      <div style="color:var(--text)">URL вказано: <b>{}</b></div>
      <div style="color:#4caf50">Локально: <b>{}</b></div>
      <div style="color:{};font-weight:700">Потрібно завантажити: {}</div>
    </div>
    <div style="background:var(--bg-card-2,#1e2d40);border:1px solid var(--border-strong);border-radius:6px;padding:10px 14px">
      <div style="font-size:11px;color:var(--text-muted);text-transform:uppercase;letter-spacing:.04em;margin-bottom:6px">🖼️ Фото товару</div>
      <div style="color:var(--text)">URL вказано: <b>{}</b></div>
      <div style="color:#4caf50">Локально: <b>{}</b></div>
      <div style="color:{};font-weight:700">Потрібно завантажити: {}</div>
    </div>
  </div>
  <button type="button" id="inv-mc-btn"
          onclick="invMediaCache()"
          style="background:#1565c0;color:#fff;border:none;border-radius:6px;padding:9px 20px;
                 font-size:13px;font-weight:600;cursor:pointer;display:inline-flex;align-items:center;gap:8px">
    📥 Завантажити відсутні файли на сервер
  </button>
  <div id="inv-mc-result" style="margin-top:10px;display:none"></div>
</div>
<script>
function invMediaCache() {{
  var btn = document.getElementById('inv-mc-btn');
  var res = document.getElementById('inv-mc-result');
  btn.disabled = true;
  btn.innerHTML = '⏳ Завантажується… (може зайняти кілька хвилин)';
  res.style.display = 'none';
  var csrf = document.cookie.match(/csrftoken=([^;]+)/);
  fetch('{}', {{
    method: 'POST',
    headers: {{'X-CSRFToken': csrf ? csrf[1] : '', 'X-Requested-With': 'XMLHttpRequest'}},
  }}).then(function(r){{ return r.json(); }}).then(function(d) {{
    btn.disabled = false;
    btn.innerHTML = '📥 Завантажити відсутні файли на сервер';
    if (!d.ok) {{ res.innerHTML = '<span style="color:var(--err)">❌ Помилка: ' + (d.error||'невідома') + '</span>'; res.style.display='block'; return; }}
    var s = d.stats;
    res.innerHTML =
      '<div style="background:var(--bg-card-2,#1e2d40);border:1px solid var(--border-strong);border-radius:6px;padding:10px 14px;max-width:460px">' +
      '<div style="font-weight:700;margin-bottom:6px;color:var(--text)">✅ Готово!</div>' +
      '<div>📄 Datasheet: <span style="color:#4caf50">+' + s.datasheets.done + ' завантажено</span>' +
           (s.datasheets.skipped ? ', ' + s.datasheets.skipped + ' вже є' : '') +
           (s.datasheets.failed  ? ', <span style="color:var(--err)">' + s.datasheets.failed + ' помилок</span>' : '') + '</div>' +
      '<div>🖼️ Фото: <span style="color:#4caf50">+' + s.images.done + ' завантажено</span>' +
           (s.images.skipped ? ', ' + s.images.skipped + ' вже є' : '') +
           (s.images.failed  ? ', <span style="color:var(--err)">' + s.images.failed + ' помилок</span>' : '') + '</div>' +
      (d.errors && d.errors.length ? '<details style="margin-top:6px"><summary style="font-size:11px;cursor:pointer;color:var(--text-muted)">' + d.errors.length + ' помилок — деталі</summary>' +
           '<pre style="font-size:10px;color:var(--err);margin:4px 0 0;overflow:auto;max-height:120px">' + d.errors.join('\\n') + '</pre></details>' : '') +
      '</div>';
    res.style.display = 'block';
  }}).catch(function(e) {{
    btn.disabled = false;
    btn.innerHTML = '📥 Завантажити відсутні файли на сервер';
    res.innerHTML = '<span style="color:var(--err)">❌ Мережева помилка: ' + e + '</span>';
    res.style.display = 'block';
  }});
}}
</script>
''',
            ds_url, ds_file,
            '#ff9800' if ds_need else '#4caf50', ds_need,
            img_url, img_file,
            '#ff9800' if img_need else '#4caf50', img_need,
            cache_url,
        )
    media_cache_panel.short_description = "Статус медіафайлів"

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path("media-cache/",
                 self.admin_site.admin_view(self.media_cache_view),
                 name="inventory_inventorysettings_media_cache"),
        ]
        return custom + urls

    def media_cache_view(self, request):
        from django.http import JsonResponse
        from django.core.files.base import ContentFile
        import urllib.request as _urllib
        import urllib.error
        import posixpath
        import os as _os

        stats  = {'datasheets': {'done': 0, 'skipped': 0, 'failed': 0},
                  'images':     {'done': 0, 'skipped': 0, 'failed': 0}}
        errors = []

        products = Product.objects.filter(is_active=True)

        for p in products:
            # ── Datasheet ──────────────────────────────────────────────
            has_ds = p.datasheet_file and p.datasheet_file.name
            if p.datasheet_url and not has_ds:
                try:
                    req = _urllib.Request(p.datasheet_url,
                                         headers={'User-Agent': 'Mozilla/5.0 (compatible; Minerva-BI/1.0)'})
                    with _urllib.urlopen(req, timeout=20) as resp:
                        data = resp.read()
                    url_path = p.datasheet_url.split('?')[0]
                    ext = _os.path.splitext(posixpath.basename(url_path))[1].lower() or '.pdf'
                    fname = p.sku.replace('/', '_').replace(' ', '_') + '_datasheet' + ext
                    p.datasheet_file.save(fname, ContentFile(data), save=True)
                    stats['datasheets']['done'] += 1
                except Exception as e:
                    stats['datasheets']['failed'] += 1
                    errors.append(f'DS {p.sku}: {e}')
            else:
                stats['datasheets']['skipped'] += 1

            # ── Image ──────────────────────────────────────────────────
            has_img = p.image and p.image.name
            if p.image_url and not has_img:
                try:
                    req = _urllib.Request(p.image_url,
                                         headers={'User-Agent': 'Mozilla/5.0 (compatible; Minerva-BI/1.0)'})
                    with _urllib.urlopen(req, timeout=15) as resp:
                        ct   = resp.headers.get('Content-Type', '')
                        data = resp.read()
                    url_path = p.image_url.split('?')[0]
                    ext = _os.path.splitext(posixpath.basename(url_path))[1].lower()
                    if not ext:
                        ext = '.png' if 'png' in ct else '.jpg'
                    fname = p.sku.replace('/', '_').replace(' ', '_') + ext
                    p.image.save(fname, ContentFile(data), save=True)
                    stats['images']['done'] += 1
                except Exception as e:
                    stats['images']['failed'] += 1
                    errors.append(f'IMG {p.sku}: {e}')
            else:
                stats['images']['skipped'] += 1

        return JsonResponse({'ok': True, 'stats': stats, 'errors': errors[:50]})

    def has_add_permission(self, request):
        return not InventorySettings.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False

    def changelist_view(self, request, extra_context=None):
        obj, _ = InventorySettings.objects.get_or_create(pk=1)
        return redirect(reverse("admin:inventory_inventorysettings_change", args=[obj.pk]))


# ── Purchase Orders ────────────────────────────────────────────────────────────

class PurchaseOrderLineInline(admin.TabularInline):
    model = PurchaseOrderLine
    extra = 0
    autocomplete_fields = ("product",)
    fields = ("product", "description", "qty_ordered",
              "qty_received", "unit_price", "currency", "notes")


@admin.register(Supplier)
class SupplierAdmin(admin.ModelAdmin):
    list_display  = ("name", "contact_person", "email", "phone", "currency", "website")
    search_fields = ("name", "contact_person", "email")
    fieldsets = (
        ("🏢 Постачальник", {
            "fields": ("name", "website", "notes")
        }),
        ("📞 Контакти", {
            "fields": ("contact_person", "email", "phone")
        }),
        ("📍 Адреса", {
            "fields": ("addr_street", ("addr_city", "addr_zip", "addr_country"))
        }),
        ("💼 Комерційні умови", {
            "fields": ("payment_terms", "currency")
        }),
    )


@admin.register(PurchaseOrder)
class PurchaseOrderAdmin(admin.ModelAdmin):
    list_display  = ("code", "supplier", "status_badge", "order_date",
                     "expected_date", "received_date", "lines_count",
                     "total_price", "tracking_number")
    list_filter   = ("supplier", "status", "order_date")
    search_fields = ("code", "tracking_number", "supplier__name", "notes")
    date_hierarchy = "order_date"
    inlines       = (PurchaseOrderLineInline,)
    readonly_fields = ("code", "created_at")
    preserve_filters = True

    def status_badge(self, obj):
        colors = {
            'draft': '#607d8b', 'ordered': '#2196f3',
            'partial': '#ff9800', 'received': '#4caf50', 'cancelled': '#f44336',
        }
        color = colors.get(obj.status, '#607d8b')
        return format_html(
            '<span style="background:{};color:#fff;padding:3px 10px;border-radius:12px;'
            'font-size:11px;font-weight:bold;white-space:nowrap">{}</span>',
            color, obj.get_status_display())
    status_badge.short_description = "Статус"
    status_badge.admin_order_field = "status"

    def lines_count(self, obj):
        c = obj.lines.count()
        return format_html('<b>{}</b> поз.', c)
    lines_count.short_description = "Позицій"


@admin.register(PurchaseOrderLine)
class PurchaseOrderLineAdmin(admin.ModelAdmin):
    list_display  = ("purchase_order", "product", "description",
                     "qty_ordered", "qty_received", "unit_price")
    list_filter   = ("purchase_order__supplier", "purchase_order__status")
    search_fields = ("purchase_order__code", "product__sku", "description")


# ── Inject inventory stats into inventory app_index context ────────────────────

def _get_inventory_stats():
    try:
        from datetime import date
        from django.db.models import Sum, OuterRef, Subquery, F, Value
        from django.db.models.functions import Coalesce
        from decimal import Decimal

        today = date.today()

        # Products below reorder point (fast subquery)
        stock_subq = (
            InventoryTransaction.objects
            .filter(product=OuterRef("pk"))
            .values("product")
            .annotate(total=Sum("qty"))
            .values("total")
        )
        critical = (
            Product.objects
            .filter(is_active=True, reorder_point__gt=0)
            .annotate(stock=Coalesce(Subquery(stock_subq), Value(Decimal("0"))))
            .filter(stock__lt=F("reorder_point"))
            .count()
        )

        # Active purchase orders (not yet fully received)
        active_po = PurchaseOrder.objects.filter(
            status__in=["draft", "ordered", "partial"]
        ).count()

        # Transactions today
        tx_today = InventoryTransaction.objects.filter(
            tx_date__date=today
        ).count()

        # Total active products
        active_products = Product.objects.filter(is_active=True).count()

        return {
            "critical":       critical,
            "active_po":      active_po,
            "tx_today":       tx_today,
            "active_products": active_products,
        }
    except Exception:
        return {"critical": "—", "active_po": "—", "tx_today": "—", "active_products": "—"}


_orig_inventory_app_index = admin.site.app_index


def _inventory_app_index(request, app_label, extra_context=None):
    if app_label == "inventory":
        extra_context = extra_context or {}
        extra_context["inventory_stats"] = _get_inventory_stats()
    return _orig_inventory_app_index(request, app_label, extra_context)


admin.site.app_index = _inventory_app_index
